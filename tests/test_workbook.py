from __future__ import annotations

from datetime import date
from decimal import Decimal
import importlib.util
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from pandas.testing import assert_frame_equal

from accounting_sim.account_mapping import DEFAULT_ACCOUNT_ROLE_MAP, build_default_account_role_mapping
from accounting_sim.canonical import (
    ACCOUNT_ROLE_MAPPING_COLUMNS,
    CHART_OF_ACCOUNTS_COLUMNS,
    ENTITY_PROFILE_COLUMNS,
    EVENT_COLUMNS,
    FISCAL_EVENT_ATTRIBUTE_COLUMNS,
    STATEMENT_MAPPING_COLUMNS,
    TAX_PARAMETER_COLUMNS,
    TAX_SCENARIO_COLUMNS,
    AccountingPeriod,
    DebitCredit,
    EventClass,
    EventDirection,
    EventNature,
    EventType,
    Origin,
    PaymentTerm,
    ScalarValueType,
    SimulationConfig,
    SchemaValidationError,
    TaxSourceType,
)
from accounting_sim.chart_of_accounts import build_default_commercial_chart, validate_chart_of_accounts
from accounting_sim.events import EVENT_SPEC_VERSION
from accounting_sim.posting import post_events
from accounting_sim.statements import (
    FINANCIAL_STATEMENT_SPEC_VERSION,
    build_default_statement_mapping,
    validate_statement_mapping,
)
from accounting_sim.tax_comparison import CBS_2026_COUNTERFACTUAL_REPORT_SPEC_VERSION
from accounting_sim.tax_context import (
    TAX_INTERFACE_SPEC_VERSION,
    TaxContext,
    build_empty_tax_context,
    validate_tax_context,
)
from accounting_sim.workbook import (
    COUNTERFACTUAL_COMPARISON_WORKBOOK_COLUMNS,
    EDITABLE_SHEETS,
    EVENT_WORKBOOK_COLUMNS,
    FISCAL_ASSESSMENT_WORKBOOK_COLUMNS,
    FISCAL_OPERATION_WORKBOOK_COLUMNS,
    TABLE_NAMES,
    WORKBOOK_SHEETS,
    WORKBOOK_SPEC_VERSION,
    WorkbookInputs,
    build_workbook,
    load_workbook_inputs,
    regenerate_workbook,
)


PERIOD = AccountingPeriod(date(2026, 1, 1), date(2026, 1, 31))
CONFIG = SimulationConfig(
    simulation_id="SIM_WORKBOOK",
    start_date=PERIOD.start_date,
    end_date=PERIOD.end_date,
    currency="BRL",
    seed=0,
    scenario_name="workbook",
    spec_version=WORKBOOK_SPEC_VERSION,
)


def chart() -> pd.DataFrame:
    return build_default_commercial_chart(PERIOD.start_date)


def event_row(
    event_id: str,
    event_date: date,
    event_type: EventType,
    amount: int,
    cost: int | None = None,
    medium: str | None = "caixa",
) -> dict[str, object]:
    direction = {
        EventType.CAPITAL_CONTRIBUTION: EventDirection.IN.value,
        EventType.PURCHASE_CASH: EventDirection.IN.value,
        EventType.SALE_CREDIT: EventDirection.OUT.value,
        EventType.CUSTOMER_RECEIPT: EventDirection.IN.value,
    }[event_type]
    nature = {
        EventType.PURCHASE_CASH: EventNature.GOOD.value,
        EventType.SALE_CREDIT: EventNature.GOOD.value,
    }.get(event_type, EventNature.FINANCIAL.value)
    payment_term = {
        EventType.PURCHASE_CASH: PaymentTerm.CASH.value,
        EventType.SALE_CREDIT: PaymentTerm.CREDIT.value,
    }.get(event_type, PaymentTerm.NA.value)
    if event_type is EventType.SALE_CREDIT:
        medium = None
    return {
        "ID_EVENTO": event_id,
        "DT_EVENTO": event_date,
        "CLASSE_EVENTO": EventClass.TRANSACTION.value,
        "TIPO_EVENTO": event_type.value,
        "DIRECAO": direction,
        "NATUREZA": nature,
        "VL_EVENTO_CENTS": amount,
        "VL_CUSTO_CENTS": cost,
        "MEIO_FINANCEIRO": medium,
        "CATEGORIA_DESPESA": None,
        "COD_PART": "PART001",
        "COND_PAGTO": payment_term,
        "DOC_REF": f"DOC-{event_id}",
        "HIST": event_type.value,
        "ORIGEM": Origin.SYNTHETIC.value,
        "SPEC_VERSION": EVENT_SPEC_VERSION,
    }


def canonical_events() -> pd.DataFrame:
    return pd.DataFrame(
        [
            event_row("E001", date(2026, 1, 1), EventType.CAPITAL_CONTRIBUTION, 10000000),
            event_row("E002", date(2026, 1, 2), EventType.PURCHASE_CASH, 3000000),
            event_row("E003", date(2026, 1, 3), EventType.SALE_CREDIT, 5000000, 2000000),
            event_row("E004", date(2026, 1, 4), EventType.CUSTOMER_RECEIPT, 3000000),
        ],
        columns=EVENT_COLUMNS,
        dtype=object,
    )


def workbook_inputs(
    chart_of_accounts: pd.DataFrame | None = None,
    account_role_mapping: pd.DataFrame | None = None,
    events: pd.DataFrame | None = None,
    statement_mapping: pd.DataFrame | None = None,
    tax_context: TaxContext | None = None,
) -> WorkbookInputs:
    chart_df = chart() if chart_of_accounts is None else chart_of_accounts
    return WorkbookInputs(
        simulation_config=CONFIG,
        chart_of_accounts=chart_df,
        account_role_mapping=build_default_account_role_mapping() if account_role_mapping is None else account_role_mapping,
        events=canonical_events() if events is None else events,
        statement_mapping=build_default_statement_mapping(chart_df) if statement_mapping is None else statement_mapping,
        tax_context=tax_context,
    )


def configured_tax_context(version: str = "VERSAO_TESTE", scenario_id: str = "CENARIO_TESTE") -> TaxContext:
    entity_profile = pd.DataFrame(
        [
            ("ENTIDADE_TESTE", "porte_teste", "micro", ScalarValueType.STRING.value, Origin.TEMPLATE.value),
            ("ENTIDADE_TESTE", "uf_teste", "RJ", ScalarValueType.STRING.value, Origin.TEMPLATE.value),
        ],
        columns=ENTITY_PROFILE_COLUMNS,
        dtype=object,
    )
    fiscal_event_attributes = pd.DataFrame(
        [
            ("E003", "atributo_fiscal_teste", "mercadoria_teste", ScalarValueType.STRING.value, Origin.TEMPLATE.value),
        ],
        columns=FISCAL_EVENT_ATTRIBUTE_COLUMNS,
        dtype=object,
    )
    tax_scenarios = pd.DataFrame(
        [
            (
                scenario_id,
                "ENTIDADE_TESTE",
                "Cenario tributario artificial",
                True,
                date(2026, 1, 31),
                "REGIME_TESTE",
                "REGIME_TESTE_IR",
                "REGIME_TESTE_CONSUMO",
                "",
                version,
                True,
            ),
        ],
        columns=TAX_SCENARIO_COLUMNS,
        dtype=object,
    )
    tax_parameters = pd.DataFrame(
        [
            (
                f"PARAM_{version}",
                version,
                "REGRA_TESTE",
                "TRIBUTO_TESTE",
                "CHAVE_TESTE",
                "123.45",
                ScalarValueType.DECIMAL.value,
                TaxSourceType.TECHNICAL.value,
                "Fonte artificial de teste",
                "https://example.invalid/fonte-teste",
                "DISPOSITIVO_TESTE",
                "NORMA_TESTE",
                date(2026, 1, 1),
                None,
                date(2026, 1, 31),
                "REGRA_TESTE_V1",
            ),
        ],
        columns=TAX_PARAMETER_COLUMNS,
        dtype=object,
    )
    return TaxContext(entity_profile, fiscal_event_attributes, tax_scenarios, tax_parameters)


def cbs_fixture_dir():
    return Path(__file__).resolve().parents[1] / "data/examples/cbs_2026"


def cbs_events() -> pd.DataFrame:
    events = pd.read_csv(cbs_fixture_dir() / "events.csv", dtype=str, keep_default_na=False)
    events["DT_EVENTO"] = pd.to_datetime(events["DT_EVENTO"]).dt.date
    events["VL_EVENTO_CENTS"] = events["VL_EVENTO_CENTS"].astype(int)
    events["VL_CUSTO_CENTS"] = events["VL_CUSTO_CENTS"].replace("", pd.NA)
    mask = events["VL_CUSTO_CENTS"].notna()
    events.loc[mask, "VL_CUSTO_CENTS"] = events.loc[mask, "VL_CUSTO_CENTS"].astype(int)
    for column in ("MEIO_FINANCEIRO", "CATEGORIA_DESPESA", "COD_PART", "DOC_REF"):
        events[column] = events[column].replace("", pd.NA)
    events.loc[events["TIPO_EVENTO"] == EventType.SALE_CREDIT.value, "MEIO_FINANCEIRO"] = pd.NA
    return events.loc[:, list(EVENT_COLUMNS)]


def cbs_tax_scenarios() -> pd.DataFrame:
    scenarios = pd.read_csv(cbs_fixture_dir() / "tax_scenarios.csv", dtype=str, keep_default_na=False)
    scenarios["DT_REFERENCIA_NORMATIVA"] = pd.to_datetime(scenarios["DT_REFERENCIA_NORMATIVA"]).dt.date
    scenarios["E_BASELINE"] = scenarios["E_BASELINE"].map(lambda value: str(value).lower() == "true")
    scenarios["ATIVO"] = scenarios["ATIVO"].map(lambda value: str(value).lower() == "true")
    return scenarios.loc[:, list(TAX_SCENARIO_COLUMNS)]


def cbs_tax_context(*, control: bool = False, invalid_control: bool = False) -> TaxContext:
    scenarios = cbs_tax_scenarios()
    if control:
        control_row = scenarios.loc[scenarios["ID_CENARIO"] == "CBS_2026_BASE"].iloc[0].copy()
        control_row["ID_CENARIO"] = "CBS_2026_CONTROLE"
        control_row["DESCRICAO"] = "controle de orquestracao"
        control_row["E_BASELINE"] = False
        control_row["ATIVO"] = True
        if invalid_control:
            control_row["REGIME_ENTIDADE"] = "simples_nacional"
        scenarios = pd.concat([scenarios, pd.DataFrame([control_row])], ignore_index=True)
    return TaxContext(
        entity_profile=pd.read_csv(cbs_fixture_dir() / "entity_profile.csv", dtype=str, keep_default_na=False),
        fiscal_event_attributes=pd.read_csv(cbs_fixture_dir() / "fiscal_event_attributes.csv", dtype=str, keep_default_na=False),
        tax_scenarios=scenarios,
        tax_parameters=pd.read_csv(cbs_fixture_dir() / "tax_parameters.csv", dtype=str, keep_default_na=False),
    )


def cbs_workbook_inputs(tax_context: TaxContext | None = None) -> WorkbookInputs:
    period = AccountingPeriod(date(2026, 8, 1), date(2026, 8, 31))
    config = SimulationConfig(
        simulation_id="SIM_CBS_WORKBOOK",
        start_date=period.start_date,
        end_date=period.end_date,
        currency="BRL",
        seed=0,
        scenario_name="cbs_workbook",
        spec_version=WORKBOOK_SPEC_VERSION,
    )
    chart_df = build_default_commercial_chart(period.start_date)
    return WorkbookInputs(
        simulation_config=config,
        chart_of_accounts=chart_df,
        account_role_mapping=build_default_account_role_mapping(),
        events=cbs_events(),
        statement_mapping=build_default_statement_mapping(chart_df),
        tax_context=tax_context if tax_context is not None else cbs_tax_context(),
    )


def build_case(tmp_path, inputs: WorkbookInputs | None = None):
    path = tmp_path / "case.xlsx"
    build_workbook(inputs or workbook_inputs(), path)
    return path, load_workbook(path, data_only=True)


def sheet_frame(path, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    table = ws.tables[TABLE_NAMES[sheet_name]]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0], dtype=object)


def sheet_text(ws) -> str:
    values = []
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value is not None:
                values.append(str(value))
    return "\n".join(values)


def cell_right_of(ws, label: str):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return ws.cell(row=cell.row, column=cell.column + 1).value
    raise AssertionError(f"Label não encontrado: {label}")


def comparison_indicator_values(ws, indicator: str) -> tuple[object, object, object]:
    for row in ws.iter_rows():
        if row[0].value == indicator:
            return row[1].value, row[2].value, row[3].value
    raise AssertionError(f"Indicador não encontrado: {indicator}")


def hyperlink_targets(ws) -> set[str]:
    return {cell.hyperlink.target for row in ws.iter_rows() for cell in row if cell.hyperlink is not None}


def presentation_values(path, sheet_name: str) -> list[list[object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def decimal_value(value: object) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.00"))


def row_by_code(frame: pd.DataFrame, code: str) -> pd.Series:
    return frame[frame["COD_CTA"] == code].iloc[0]


def row_by_line(frame: pd.DataFrame, code: str) -> pd.Series:
    return frame[frame["COD_LINHA"] == code].iloc[0]


def test_openpyxl_is_available_as_dependency():
    assert importlib.util.find_spec("openpyxl") is not None


def test_build_workbook_creates_xlsx_and_can_be_reopened(tmp_path):
    path, wb = build_case(tmp_path)
    assert path.exists()
    assert path.suffix == ".xlsx"
    assert "README" in wb.sheetnames


def test_workbook_sheets_are_exactly_in_canonical_order(tmp_path):
    _, wb = build_case(tmp_path)
    assert WORKBOOK_SHEETS == (
        "RESUMO",
        "ENTRADAS",
        "COMPARACAO",
        "README",
        "CONFIG",
        "ENTIDADE",
        "PLANO_CONTAS",
        "MAPEAMENTO_CONTAS",
        "EVENTOS",
        "EVENTOS_FISCAIS",
        "LANCAMENTOS",
        "PARTIDAS",
        "VINCULO_EVENTO_LCTO",
        "DIARIO",
        "RAZAO",
        "BALANCETE",
        "MAPEAMENTO_DF",
        "BP",
        "DRE",
        "CENARIOS_TRIBUTARIOS",
        "FISCAL_PARAM",
        "FISCAL_RESULTADOS_OPERACAO",
        "FISCAL_APURACAO",
        "COMPARATIVO_CENARIOS",
        "VALIDACOES",
        "PROVENIENCIA",
    )
    assert tuple(wb.sheetnames) == WORKBOOK_SHEETS


def test_tax_output_sheets_are_derived_not_editable(tmp_path):
    _, wb = build_case(tmp_path)
    forbidden = {
        "CENTROS_CUSTO",
        "PARTICIPANTES",
        "HISTORICOS",
        "DFC",
        "DVA",
    }
    assert forbidden.isdisjoint(set(wb.sheetnames))
    assert {
        "RESUMO",
        "ENTRADAS",
        "COMPARACAO",
        "FISCAL_RESULTADOS_OPERACAO",
        "FISCAL_APURACAO",
        "COMPARATIVO_CENARIOS",
    }.issubset(set(wb.sheetnames))
    assert {
        "FISCAL_RESULTADOS_OPERACAO",
        "FISCAL_APURACAO",
        "COMPARATIVO_CENARIOS",
    }.isdisjoint(EDITABLE_SHEETS)


def test_presentation_sheets_are_first_and_not_editable(tmp_path):
    _, wb = build_case(tmp_path)

    assert wb.sheetnames[:3] == ["RESUMO", "ENTRADAS", "COMPARACAO"]
    assert {"RESUMO", "ENTRADAS", "COMPARACAO"}.isdisjoint(EDITABLE_SHEETS)


def test_summary_sheet_explains_canonical_case_without_technical_ids(tmp_path):
    path, wb = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    ws = wb["RESUMO"]
    text = sheet_text(ws)

    assert "2026-08-01 a 2026-08-31" in text
    assert "OPERAÇÕES DO PERÍODO" in text
    assert "Compra de mercadoria à vista" in text
    assert "Venda a prazo" in text
    assert "Aporte de capital" in text
    assert decimal_value(cell_right_of(ws, "Compra de mercadoria à vista")) == Decimal("1010.00")
    assert decimal_value(cell_right_of(ws, "Venda a prazo")) == Decimal("2000.00")
    assert decimal_value(cell_right_of(ws, "Aporte de capital")) == Decimal("5000.00")

    assert "RESULTADO CONTÁBIL" in text
    assert decimal_value(cell_right_of(ws, "Receita")) == Decimal("2000.00")
    assert decimal_value(cell_right_of(ws, "CMV")) == Decimal("-1000.00")
    assert decimal_value(cell_right_of(ws, "Resultado do período")) == Decimal("1000.00")
    assert decimal_value(cell_right_of(ws, "Ativo final")) == Decimal("6000.00")
    assert decimal_value(cell_right_of(ws, "Passivo")) == Decimal("0.00")
    assert decimal_value(cell_right_of(ws, "Patrimônio líquido")) == Decimal("6000.00")

    assert "RESULTADO TRIBUTÁRIO - BASELINE" in text
    assert cell_right_of(ws, "Tributo") == "CBS"
    assert decimal_value(cell_right_of(ws, "Débitos")) == Decimal("18.00")
    assert decimal_value(cell_right_of(ws, "Créditos")) == Decimal("9.00")
    assert decimal_value(cell_right_of(ws, "CBS apurada")) == Decimal("9.00")
    assert decimal_value(cell_right_of(ws, "CBS a recolher")) == Decimal("0.00")
    assert cell_right_of(ws, "Impacto em caixa") == "não calculado"
    assert cell_right_of(ws, "Impacto em DRE") == "não calculado"
    assert "CBS_2026_CONTROLE é um controle estrutural" in text

    assert sheet_frame(path, "FISCAL_APURACAO").loc[0, "P_CASH"] is None


def test_inputs_sheet_guides_editable_sheets_by_care_level(tmp_path):
    _, wb = build_case(tmp_path)
    ws = wb["ENTRADAS"]
    text = sheet_text(ws)

    assert "ENTRADAS DE USO COMUM" in text
    assert "CONFIGURAÇÕES ESTRUTURAIS" in text
    assert "PARÂMETROS NORMATIVOS - ALTA SENSIBILIDADE" in text
    for sheet_name in EDITABLE_SHEETS:
        assert sheet_name in text
        assert f"#'{sheet_name}'!A1" in hyperlink_targets(ws)
    assert "muito alto" in text


def test_comparison_sheet_presents_counterfactual_values_without_ranking(tmp_path):
    path, wb = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    ws = wb["COMPARACAO"]
    text = sheet_text(ws)
    assessment = sheet_frame(path, "FISCAL_APURACAO")
    comparison = sheet_frame(path, "COMPARATIVO_CENARIOS").iloc[0]
    baseline = assessment.loc[assessment["ID_CENARIO"] == "CBS_2026_BASE"].iloc[0]
    alternative = assessment.loc[assessment["ID_CENARIO"] == "CBS_2026_CONTROLE"].iloc[0]

    assert cell_right_of(ws, "Cenário baseline") == "CBS_2026_BASE"
    assert cell_right_of(ws, "Cenário alternativo") == "CBS_2026_CONTROLE"
    assert cell_right_of(ws, "Tributo") == "CBS"
    for indicator in ("Saldo apurado", "Valor a recolher", "Saldo credor", "Impacto em caixa", "Impacto em DRE"):
        assert indicator in text

    base_value, alternative_value, delta = comparison_indicator_values(ws, "Saldo apurado")
    assert decimal_value(base_value) == decimal_value(baseline["S_APUR"])
    assert decimal_value(alternative_value) == decimal_value(alternative["S_APUR"])
    assert decimal_value(delta) == decimal_value(comparison["DELTA_S_APUR"])

    base_value, alternative_value, delta = comparison_indicator_values(ws, "Valor a recolher")
    assert decimal_value(base_value) == decimal_value(baseline["T_RECOLHER"])
    assert decimal_value(alternative_value) == decimal_value(alternative["T_RECOLHER"])
    assert decimal_value(delta) == decimal_value(comparison["DELTA_T_RECOLHER"])

    assert comparison_indicator_values(ws, "Impacto em caixa") == ("não calculado", "não calculado", "não calculado")
    assert comparison_indicator_values(ws, "Impacto em DRE") == ("não calculado", "não calculado", "não calculado")
    assert "melhor cenário" not in text.lower()
    assert "ranking" not in text.lower()
    assert "score" not in text.lower()
    assert "controle estrutural" in text


def test_presentation_internal_hyperlinks_target_existing_sheets(tmp_path):
    _, wb = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    expected = {
        "RESUMO": {"#'ENTRADAS'!A1", "#'COMPARACAO'!A1", "#'BP'!A1", "#'DRE'!A1", "#'FISCAL_APURACAO'!A1"},
        "COMPARACAO": {"#'FISCAL_APURACAO'!A1", "#'COMPARATIVO_CENARIOS'!A1", "#'CENARIOS_TRIBUTARIOS'!A1"},
    }
    for sheet_name, targets in expected.items():
        assert targets.issubset(hyperlink_targets(wb[sheet_name]))
    for target in hyperlink_targets(wb["ENTRADAS"]):
        assert target.startswith("#'")
        assert target.endswith("'!A1")
        assert target.removeprefix("#'").removesuffix("'!A1") in wb.sheetnames


def test_regenerate_rebuilds_presentation_sheets_and_discards_manual_edits(tmp_path):
    path, _ = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    wb = load_workbook(path)
    wb["RESUMO"]["A1"] = "ADULTERADO"
    wb["ENTRADAS"]["A1"] = "ADULTERADO"
    wb["COMPARACAO"]["A1"] = "ADULTERADO"
    wb.save(path)

    regenerated = tmp_path / "regenerated_presentation.xlsx"
    regenerate_workbook(path, regenerated)
    rebuilt = load_workbook(regenerated, data_only=True)

    assert rebuilt["RESUMO"]["A1"].value == "RESUMO"
    assert rebuilt["ENTRADAS"]["A1"].value == "ENTRADAS"
    assert rebuilt["COMPARACAO"]["A1"].value == "COMPARAÇÃO"
    assert decimal_value(cell_right_of(rebuilt["COMPARACAO"], "Saldo apurado")) == Decimal("9.00")


def test_all_previous_technical_sheets_remain_available_after_presentation_layer(tmp_path):
    _, wb = build_case(tmp_path)
    previous_technical_sheets = {
        "README",
        "CONFIG",
        "ENTIDADE",
        "PLANO_CONTAS",
        "MAPEAMENTO_CONTAS",
        "EVENTOS",
        "EVENTOS_FISCAIS",
        "LANCAMENTOS",
        "PARTIDAS",
        "VINCULO_EVENTO_LCTO",
        "DIARIO",
        "RAZAO",
        "BALANCETE",
        "MAPEAMENTO_DF",
        "BP",
        "DRE",
        "CENARIOS_TRIBUTARIOS",
        "FISCAL_PARAM",
        "FISCAL_RESULTADOS_OPERACAO",
        "FISCAL_APURACAO",
        "COMPARATIVO_CENARIOS",
        "VALIDACOES",
        "PROVENIENCIA",
    }
    assert previous_technical_sheets.issubset(set(wb.sheetnames))


def test_presentation_sheets_are_deterministic(tmp_path):
    first = tmp_path / "first_presentation.xlsx"
    second = tmp_path / "second_presentation.xlsx"
    inputs = cbs_workbook_inputs(cbs_tax_context(control=True))
    build_workbook(inputs, first)
    build_workbook(inputs, second)

    for sheet_name in ("RESUMO", "ENTRADAS", "COMPARACAO"):
        assert presentation_values(first, sheet_name) == presentation_values(second, sheet_name)


def test_named_tables_exist_and_cover_expected_row_counts(tmp_path):
    path, wb = build_case(tmp_path)
    expected_rows = {
        "CONFIG": 7,
        "ENTIDADE": 0,
        "PLANO_CONTAS": len(chart()),
        "MAPEAMENTO_CONTAS": len(build_default_account_role_mapping()),
        "EVENTOS": len(canonical_events()),
        "EVENTOS_FISCAIS": 0,
        "LANCAMENTOS": 5,
        "PARTIDAS": 10,
        "VINCULO_EVENTO_LCTO": 5,
        "DIARIO": 10,
        "RAZAO": 10,
        "BALANCETE": len(chart()[chart()["IND_CTA"] == "A"]),
        "MAPEAMENTO_DF": len(build_default_statement_mapping(chart())),
        "BP": 22,
        "DRE": 11,
        "CENARIOS_TRIBUTARIOS": 0,
        "FISCAL_PARAM": 0,
        "FISCAL_RESULTADOS_OPERACAO": 0,
        "FISCAL_APURACAO": 0,
        "COMPARATIVO_CENARIOS": 0,
        "VALIDACOES": 12,
        "PROVENIENCIA": 14,
    }
    for sheet_name, table_name in TABLE_NAMES.items():
        assert table_name in wb[sheet_name].tables
        frame = sheet_frame(path, sheet_name)
        assert len(frame) == expected_rows[sheet_name]


def test_tabular_sheets_have_freeze_panes_and_filters(tmp_path):
    _, wb = build_case(tmp_path)
    for sheet_name, table_name in TABLE_NAMES.items():
        ws = wb[sheet_name]
        assert ws.freeze_panes == "A2"
        assert ws.tables[table_name].autoFilter is not None


def test_config_round_trip_preserves_types(tmp_path):
    path, _ = build_case(tmp_path)
    reloaded = load_workbook_inputs(path)
    assert reloaded.simulation_config == CONFIG
    assert isinstance(reloaded.simulation_config.start_date, date)
    assert isinstance(reloaded.simulation_config.seed, int)


def test_chart_round_trip_preserves_keys_dates_bools_and_hierarchy(tmp_path):
    path, _ = build_case(tmp_path)
    reloaded = load_workbook_inputs(path)
    assert tuple(reloaded.chart_of_accounts.columns) == CHART_OF_ACCOUNTS_COLUMNS
    assert list(reloaded.chart_of_accounts["COD_CTA"]) == list(chart()["COD_CTA"])
    assert reloaded.chart_of_accounts["DT_ALT"].map(type).eq(date).all()
    assert reloaded.chart_of_accounts["ATIVA"].map(type).eq(bool).all()
    assert validate_chart_of_accounts(reloaded.chart_of_accounts).ok is True


def test_mapping_round_trip_preserves_roles_and_codes(tmp_path):
    path, _ = build_case(tmp_path)
    reloaded = load_workbook_inputs(path)
    assert tuple(reloaded.account_role_mapping.columns) == ACCOUNT_ROLE_MAPPING_COLUMNS
    assert_frame_equal(reloaded.account_role_mapping, build_default_account_role_mapping(), check_dtype=False)


def test_statement_mapping_round_trip_preserves_accounts_and_lines(tmp_path):
    path, _ = build_case(tmp_path)
    reloaded = load_workbook_inputs(path)
    expected = build_default_statement_mapping(chart())
    assert tuple(reloaded.statement_mapping.columns) == STATEMENT_MAPPING_COLUMNS
    assert_frame_equal(reloaded.statement_mapping, expected, check_dtype=False)
    assert validate_statement_mapping(reloaded.statement_mapping, reloaded.chart_of_accounts).ok is True


def test_empty_tax_context_round_trip_preserves_exact_schemas(tmp_path):
    path, _ = build_case(tmp_path)
    reloaded = load_workbook_inputs(path)
    expected = build_empty_tax_context()
    assert reloaded.tax_context is not None
    assert tuple(reloaded.tax_context.entity_profile.columns) == ENTITY_PROFILE_COLUMNS
    assert tuple(reloaded.tax_context.fiscal_event_attributes.columns) == FISCAL_EVENT_ATTRIBUTE_COLUMNS
    assert tuple(reloaded.tax_context.tax_scenarios.columns) == TAX_SCENARIO_COLUMNS
    assert tuple(reloaded.tax_context.tax_parameters.columns) == TAX_PARAMETER_COLUMNS
    assert reloaded.tax_context.entity_profile.empty
    assert reloaded.tax_context.fiscal_event_attributes.empty
    assert reloaded.tax_context.tax_scenarios.empty
    assert reloaded.tax_context.tax_parameters.empty
    assert validate_tax_context(reloaded.tax_context, reloaded.events).ok is True
    assert_frame_equal(reloaded.tax_context.entity_profile, expected.entity_profile, check_dtype=False)


def test_configured_tax_context_round_trip_preserves_values_dates_and_bools(tmp_path):
    context = configured_tax_context()
    path, _ = build_case(tmp_path, workbook_inputs(tax_context=context))
    reloaded = load_workbook_inputs(path)
    assert reloaded.tax_context is not None
    assert_frame_equal(reloaded.tax_context.entity_profile, context.entity_profile, check_dtype=False)
    assert_frame_equal(reloaded.tax_context.fiscal_event_attributes, context.fiscal_event_attributes, check_dtype=False)
    assert_frame_equal(reloaded.tax_context.tax_scenarios, context.tax_scenarios, check_dtype=False)
    assert_frame_equal(reloaded.tax_context.tax_parameters, context.tax_parameters, check_dtype=False)
    assert isinstance(reloaded.tax_context.tax_scenarios.loc[0, "DT_REFERENCIA_NORMATIVA"], date)
    assert isinstance(reloaded.tax_context.tax_scenarios.loc[0, "E_BASELINE"], bool)
    assert isinstance(reloaded.tax_context.tax_scenarios.loc[0, "ATIVO"], bool)
    assert isinstance(reloaded.tax_context.tax_parameters.loc[0, "VIG_INI"], date)
    assert isinstance(reloaded.tax_context.tax_parameters.loc[0, "DATA_CONSULTA"], date)
    assert reloaded.tax_context.tax_parameters.loc[0, "VALOR"] == "123.45"
    assert validate_tax_context(reloaded.tax_context, reloaded.events).ok is True


def test_events_round_trip_preserves_ids_dates_and_cents(tmp_path):
    path, _ = build_case(tmp_path)
    reloaded = load_workbook_inputs(path)
    assert tuple(reloaded.events.columns) == EVENT_COLUMNS
    assert list(reloaded.events["ID_EVENTO"]) == list(canonical_events()["ID_EVENTO"])
    assert list(reloaded.events["DT_EVENTO"]) == list(canonical_events()["DT_EVENTO"])
    assert list(reloaded.events["VL_EVENTO_CENTS"]) == list(canonical_events()["VL_EVENTO_CENTS"])
    assert list(reloaded.events["VL_CUSTO_CENTS"]) == list(canonical_events()["VL_CUSTO_CENTS"])


def test_events_sheet_uses_reais_columns_not_cents(tmp_path):
    path, _ = build_case(tmp_path)
    events = sheet_frame(path, "EVENTOS")
    assert tuple(events.columns) == EVENT_WORKBOOK_COLUMNS
    assert "VL_EVENTO_CENTS" not in events.columns
    assert decimal_value(events.loc[events["ID_EVENTO"] == "E001", "VL_EVENTO"].iloc[0]) == Decimal("100000.00")


def test_event_columns_are_not_changed_by_tax_interface(tmp_path):
    path, _ = build_case(tmp_path, workbook_inputs(tax_context=configured_tax_context()))
    reloaded = load_workbook_inputs(path)
    assert tuple(EVENT_COLUMNS) == (
        "ID_EVENTO",
        "DT_EVENTO",
        "CLASSE_EVENTO",
        "TIPO_EVENTO",
        "DIRECAO",
        "NATUREZA",
        "VL_EVENTO_CENTS",
        "VL_CUSTO_CENTS",
        "MEIO_FINANCEIRO",
        "CATEGORIA_DESPESA",
        "COD_PART",
        "COND_PAGTO",
        "DOC_REF",
        "HIST",
        "ORIGEM",
        "SPEC_VERSION",
    )
    assert_frame_equal(reloaded.events, canonical_events(), check_dtype=False)


def test_money_reader_rejects_more_than_two_decimal_places(tmp_path):
    path, _ = build_case(tmp_path)
    wb = load_workbook(path)
    ws = wb["EVENTOS"]
    headers = [cell.value for cell in ws[1]]
    ws.cell(row=2, column=headers.index("VL_EVENTO") + 1, value="100000.001")
    wb.save(path)
    try:
        load_workbook_inputs(path)
    except SchemaValidationError as exc:
        assert "duas casas" in str(exc)
    else:
        raise AssertionError("Valor monetário com mais de duas casas deveria ser rejeitado.")


def test_canonical_workbook_materializes_five_entries_and_ten_postings(tmp_path):
    path, _ = build_case(tmp_path)
    assert len(sheet_frame(path, "LANCAMENTOS")) == 5
    assert len(sheet_frame(path, "PARTIDAS")) == 10


def test_canonical_workbook_debit_and_credit_totals_are_brl(tmp_path):
    path, _ = build_case(tmp_path)
    postings = sheet_frame(path, "PARTIDAS")
    debit_total = sum(decimal_value(value) for value in postings.loc[postings["IND_DC"] == "D", "VL_DC"])
    credit_total = sum(decimal_value(value) for value in postings.loc[postings["IND_DC"] == "C", "VL_DC"])
    assert debit_total == Decimal("230000.00")
    assert credit_total == Decimal("230000.00")


def test_ledger_sheet_presents_expected_canonical_balances(tmp_path):
    path, _ = build_case(tmp_path)
    ledger = sheet_frame(path, "RAZAO")
    last_by_account = ledger.groupby("COD_CTA", sort=False).tail(1).set_index("COD_CTA")
    assert decimal_value(last_by_account.loc["1.1.01.01", "SALDO_ABS"]) == Decimal("100000.00")
    assert decimal_value(last_by_account.loc["1.1.02.01", "SALDO_ABS"]) == Decimal("20000.00")
    assert decimal_value(last_by_account.loc["1.1.03.01", "SALDO_ABS"]) == Decimal("10000.00")
    assert decimal_value(last_by_account.loc["3.1.01.01", "SALDO_ABS"]) == Decimal("100000.00")
    assert decimal_value(last_by_account.loc["4.1.01.01", "SALDO_ABS"]) == Decimal("50000.00")
    assert decimal_value(last_by_account.loc["4.2.01.01", "SALDO_ABS"]) == Decimal("20000.00")
    assert last_by_account.loc["3.1.01.01", "IND_DC_SALDO"] == DebitCredit.CREDIT.value


def test_trial_balance_presents_expected_debit_and_credit_balances_in_brl(tmp_path):
    path, _ = build_case(tmp_path)
    trial_balance = sheet_frame(path, "BALANCETE")
    debit_total = sum(decimal_value(value) for value in trial_balance.loc[trial_balance["IND_DC_FIN"] == "D", "VL_SLD_FIN"])
    credit_total = sum(decimal_value(value) for value in trial_balance.loc[trial_balance["IND_DC_FIN"] == "C", "VL_SLD_FIN"])
    assert debit_total == Decimal("150000.00")
    assert credit_total == Decimal("150000.00")


def test_income_statement_sheet_presents_canonical_result_in_brl(tmp_path):
    path, _ = build_case(tmp_path)
    dre = sheet_frame(path, "DRE")
    assert decimal_value(row_by_line(dre, "DRE_RECEITA_VENDAS")["VL"]) == Decimal("50000.00")
    assert decimal_value(row_by_line(dre, "DRE_CMV")["VL"]) == Decimal("-20000.00")
    assert decimal_value(row_by_line(dre, "DRE_RESULTADO_BRUTO")["VL"]) == Decimal("30000.00")
    assert decimal_value(row_by_line(dre, "DRE_RESULTADO_PERIODO")["VL"]) == Decimal("30000.00")


def test_balance_sheet_sheet_presents_canonical_totals_in_brl(tmp_path):
    path, _ = build_case(tmp_path)
    bp = sheet_frame(path, "BP")
    assert decimal_value(row_by_line(bp, "BP_ATIVO")["VL"]) == Decimal("130000.00")
    assert decimal_value(row_by_line(bp, "BP_CAPITAL")["VL"]) == Decimal("100000.00")
    assert decimal_value(row_by_line(bp, "BP_RESULTADO_PERIODO")["VL"]) == Decimal("30000.00")
    assert decimal_value(row_by_line(bp, "BP_PATRIMONIO_LIQUIDO")["VL"]) == Decimal("130000.00")
    assert decimal_value(row_by_line(bp, "BP_TOTAL_PASSIVO_PL")["VL"]) == Decimal("130000.00")


def test_workbook_preserves_previous_core_invariants_after_adding_statements(tmp_path):
    path, _ = build_case(tmp_path)
    postings = sheet_frame(path, "PARTIDAS")
    trial_balance = sheet_frame(path, "BALANCETE")
    assert len(sheet_frame(path, "LANCAMENTOS")) == 5
    assert len(postings) == 10
    assert sum(decimal_value(value) for value in postings.loc[postings["IND_DC"] == "D", "VL_DC"]) == Decimal("230000.00")
    assert sum(decimal_value(value) for value in postings.loc[postings["IND_DC"] == "C", "VL_DC"]) == Decimal("230000.00")
    assert sum(decimal_value(value) for value in trial_balance.loc[trial_balance["IND_DC_FIN"] == "D", "VL_SLD_FIN"]) == Decimal("150000.00")
    assert sum(decimal_value(value) for value in trial_balance.loc[trial_balance["IND_DC_FIN"] == "C", "VL_SLD_FIN"]) == Decimal("150000.00")


def test_validations_sheet_has_no_failures_for_canonical_case(tmp_path):
    path, _ = build_case(tmp_path)
    validations = sheet_frame(path, "VALIDACOES")
    assert set(validations["OK"]) == {True}
    assert set(validations["MENSAGEM"]) == {"ok"}
    assert {
        "ENTIDADE",
        "EVENTOS_FISCAIS",
        "CENARIOS_TRIBUTARIOS",
        "FISCAL_PARAM",
        "CONTEXTO_TRIBUTARIO",
    }.issubset(set(validations["ETAPA"]))


def test_validations_sheet_has_no_failures_for_configured_tax_context(tmp_path):
    path, _ = build_case(tmp_path, workbook_inputs(tax_context=configured_tax_context()))
    validations = sheet_frame(path, "VALIDACOES")
    assert set(validations["OK"]) == {True}
    assert {
        "ENTIDADE",
        "EVENTOS_FISCAIS",
        "CENARIOS_TRIBUTARIOS",
        "FISCAL_PARAM",
        "CONTEXTO_TRIBUTARIO",
    }.issubset(set(validations["ETAPA"]))


def test_provenance_contains_spec_and_rule_versions(tmp_path):
    path, _ = build_case(tmp_path)
    provenance = sheet_frame(path, "PROVENIENCIA").set_index("CHAVE")["VALOR"].to_dict()
    assert provenance["workbook_spec_version"] == WORKBOOK_SPEC_VERSION
    assert provenance["financial_statement_spec_version"] == FINANCIAL_STATEMENT_SPEC_VERSION
    assert provenance["tax_interface_spec_version"] == TAX_INTERFACE_SPEC_VERSION
    assert provenance["counterfactual_report_spec_version"] == CBS_2026_COUNTERFACTUAL_REPORT_SPEC_VERSION
    assert provenance["posting_rule_version"] == "posting_rules_v1"
    assert provenance["statement_mapping_source"] == "MAPEAMENTO_DF"
    assert provenance["tax_context_configured"] == "FALSE"
    assert provenance["tax_normative_versions"] is None
    assert provenance["simulation_id"] == CONFIG.simulation_id


def test_provenance_records_configured_tax_context_without_copying_parameters(tmp_path):
    path, _ = build_case(tmp_path, workbook_inputs(tax_context=configured_tax_context(version="VERSAO_TESTE")))
    provenance_frame = sheet_frame(path, "PROVENIENCIA")
    provenance = provenance_frame.set_index("CHAVE")["VALOR"].to_dict()
    assert provenance["tax_context_configured"] == "TRUE"
    assert provenance["tax_normative_versions"] == "VERSAO_TESTE"
    assert "TRIBUTO_TESTE" not in set(provenance_frame["VALOR"])


def test_manual_trial_balance_edit_is_ignored_by_regeneration(tmp_path):
    path, _ = build_case(tmp_path)
    wb = load_workbook(path)
    ws = wb["BALANCETE"]
    headers = [cell.value for cell in ws[1]]
    value_col = headers.index("VL_SLD_FIN") + 1
    code_col = headers.index("COD_CTA") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=code_col).value == "1.1.01.01":
            ws.cell(row=row, column=value_col, value=999999.99)
            break
    wb.save(path)

    regenerated = tmp_path / "regenerated_trial.xlsx"
    regenerate_workbook(path, regenerated)
    trial_balance = sheet_frame(regenerated, "BALANCETE")
    assert decimal_value(row_by_code(trial_balance, "1.1.01.01")["VL_SLD_FIN"]) == Decimal("100000.00")


def test_manual_ledger_edit_is_ignored_by_regeneration(tmp_path):
    path, _ = build_case(tmp_path)
    wb = load_workbook(path)
    ws = wb["RAZAO"]
    headers = [cell.value for cell in ws[1]]
    value_col = headers.index("SALDO_ABS") + 1
    code_col = headers.index("COD_CTA") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=code_col).value == "1.1.01.01":
            ws.cell(row=row, column=value_col, value=999999.99)
    wb.save(path)

    regenerated = tmp_path / "regenerated_ledger.xlsx"
    regenerate_workbook(path, regenerated)
    ledger = sheet_frame(regenerated, "RAZAO")
    caixa_final = ledger[ledger["COD_CTA"] == "1.1.01.01"].tail(1).iloc[0]
    assert decimal_value(caixa_final["SALDO_ABS"]) == Decimal("100000.00")


def test_manual_balance_sheet_edit_is_ignored_by_regeneration(tmp_path):
    path, _ = build_case(tmp_path)
    wb = load_workbook(path)
    ws = wb["BP"]
    headers = [cell.value for cell in ws[1]]
    value_col = headers.index("VL") + 1
    line_col = headers.index("COD_LINHA") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=line_col).value == "BP_ATIVO":
            ws.cell(row=row, column=value_col, value=999999.99)
            break
    wb.save(path)

    regenerated = tmp_path / "regenerated_bp.xlsx"
    regenerate_workbook(path, regenerated)
    bp = sheet_frame(regenerated, "BP")
    assert decimal_value(row_by_line(bp, "BP_ATIVO")["VL"]) == Decimal("130000.00")


def test_manual_income_statement_edit_is_ignored_by_regeneration(tmp_path):
    path, _ = build_case(tmp_path)
    wb = load_workbook(path)
    ws = wb["DRE"]
    headers = [cell.value for cell in ws[1]]
    value_col = headers.index("VL") + 1
    line_col = headers.index("COD_LINHA") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=line_col).value == "DRE_RESULTADO_PERIODO":
            ws.cell(row=row, column=value_col, value=999999.99)
            break
    wb.save(path)

    regenerated = tmp_path / "regenerated_dre.xlsx"
    regenerate_workbook(path, regenerated)
    dre = sheet_frame(regenerated, "DRE")
    assert decimal_value(row_by_line(dre, "DRE_RESULTADO_PERIODO")["VL"]) == Decimal("30000.00")


def test_changing_only_tax_scenarios_and_parameters_does_not_change_accounting_core(tmp_path):
    first = tmp_path / "first_tax_context.xlsx"
    second = tmp_path / "second_tax_context.xlsx"
    build_workbook(workbook_inputs(tax_context=configured_tax_context(version="VERSAO_TESTE_A", scenario_id="CENARIO_A")), first)
    build_workbook(workbook_inputs(tax_context=configured_tax_context(version="VERSAO_TESTE_B", scenario_id="CENARIO_B")), second)
    for sheet_name in (
        "LANCAMENTOS",
        "PARTIDAS",
        "VINCULO_EVENTO_LCTO",
        "DIARIO",
        "RAZAO",
        "BALANCETE",
        "BP",
        "DRE",
    ):
        assert_frame_equal(sheet_frame(first, sheet_name), sheet_frame(second, sheet_name), check_dtype=False)


def test_one_active_tax_scenario_materializes_empty_tax_outputs(tmp_path):
    path, wb = build_case(tmp_path, workbook_inputs(tax_context=configured_tax_context()))

    assert {
        "FISCAL_RESULTADOS_OPERACAO",
        "FISCAL_APURACAO",
        "COMPARATIVO_CENARIOS",
    }.issubset(set(wb.sheetnames))
    assert sheet_frame(path, "FISCAL_RESULTADOS_OPERACAO").empty
    assert sheet_frame(path, "FISCAL_APURACAO").empty
    assert sheet_frame(path, "COMPARATIVO_CENARIOS").empty


def test_empty_tax_context_materializes_empty_tax_output_headers(tmp_path):
    path, _ = build_case(tmp_path)

    assert tuple(sheet_frame(path, "FISCAL_RESULTADOS_OPERACAO").columns) == FISCAL_OPERATION_WORKBOOK_COLUMNS
    assert tuple(sheet_frame(path, "FISCAL_APURACAO").columns) == FISCAL_ASSESSMENT_WORKBOOK_COLUMNS
    assert tuple(sheet_frame(path, "COMPARATIVO_CENARIOS").columns) == COUNTERFACTUAL_COMPARISON_WORKBOOK_COLUMNS
    assert sheet_frame(path, "FISCAL_RESULTADOS_OPERACAO").empty
    assert sheet_frame(path, "FISCAL_APURACAO").empty
    assert sheet_frame(path, "COMPARATIVO_CENARIOS").empty


def test_valid_counterfactual_tax_context_populates_fiscal_output_sheets(tmp_path):
    path, _ = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))

    operations = sheet_frame(path, "FISCAL_RESULTADOS_OPERACAO")
    assessment = sheet_frame(path, "FISCAL_APURACAO")
    comparison = sheet_frame(path, "COMPARATIVO_CENARIOS")

    assert len(operations) == 4
    assert len(assessment) == 2
    assert len(comparison) == 1
    assert set(operations["ID_CENARIO"]) == {"CBS_2026_BASE", "CBS_2026_CONTROLE"}
    assert list(assessment["ID_CENARIO"]) == ["CBS_2026_BASE", "CBS_2026_CONTROLE"]
    assert comparison.iloc[0]["ID_CENARIO_BASE"] == "CBS_2026_BASE"
    assert comparison.iloc[0]["ID_CENARIO"] == "CBS_2026_CONTROLE"


def test_fiscal_outputs_are_presented_in_reais_and_preserve_unknowns_as_blank(tmp_path):
    path, _ = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    operations = sheet_frame(path, "FISCAL_RESULTADOS_OPERACAO")
    assessment = sheet_frame(path, "FISCAL_APURACAO")
    comparison = sheet_frame(path, "COMPARATIVO_CENARIOS")

    purchase = operations.loc[
        (operations["ID_CENARIO"] == "CBS_2026_BASE") & (operations["ID_EVENTO"] == "E001")
    ].iloc[0]
    sale = operations.loc[
        (operations["ID_CENARIO"] == "CBS_2026_BASE") & (operations["ID_EVENTO"] == "E002")
    ].iloc[0]
    assert decimal_value(purchase["BASE"]) == Decimal("1000.00")
    assert decimal_value(purchase["CREDITO"]) == Decimal("9.00")
    assert decimal_value(sale["DEBITO"]) == Decimal("18.00")

    base_assessment = assessment.loc[assessment["ID_CENARIO"] == "CBS_2026_BASE"].iloc[0]
    assert decimal_value(base_assessment["S_APUR"]) == Decimal("9.00")
    assert decimal_value(base_assessment["T_RECOLHER"]) == Decimal("0.00")
    assert base_assessment["P_CASH"] is None
    assert base_assessment["E_DRE"] is None
    assert decimal_value(base_assessment["C_SALDO"]) == Decimal("0.00")

    row = comparison.iloc[0]
    assert decimal_value(row["DELTA_S_APUR"]) == Decimal("0.00")
    assert decimal_value(row["DELTA_T_RECOLHER"]) == Decimal("0.00")
    assert row["DELTA_P_CASH"] is None
    assert row["DELTA_E_DRE"] is None
    assert decimal_value(row["DELTA_C_SALDO"]) == Decimal("0.00")


def test_fiscal_rate_is_stored_as_fraction_and_formatted_as_percent(tmp_path):
    path, wb = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    operations = sheet_frame(path, "FISCAL_RESULTADOS_OPERACAO")
    assert Decimal(str(operations.iloc[0]["ALIQUOTA"])) == Decimal("0.009")

    ws = wb["FISCAL_RESULTADOS_OPERACAO"]
    headers = [cell.value for cell in ws[1]]
    rate_col = headers.index("ALIQUOTA") + 1
    assert ws.cell(row=2, column=rate_col).number_format == "0.0000%"


def test_regenerate_recalculates_fiscal_outputs_and_discards_manual_edits(tmp_path):
    path, _ = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    wb = load_workbook(path)
    ws = wb["FISCAL_APURACAO"]
    headers = [cell.value for cell in ws[1]]
    ws.cell(row=2, column=headers.index("S_APUR") + 1, value=999999.99)
    comparison_ws = wb["COMPARATIVO_CENARIOS"]
    comparison_headers = [cell.value for cell in comparison_ws[1]]
    comparison_ws.cell(row=2, column=comparison_headers.index("DELTA_S_APUR") + 1, value=12345.67)
    wb.save(path)

    regenerated = tmp_path / "regenerated_tax_outputs.xlsx"
    regenerate_workbook(path, regenerated)

    assessment = sheet_frame(regenerated, "FISCAL_APURACAO")
    comparison = sheet_frame(regenerated, "COMPARATIVO_CENARIOS")
    assert decimal_value(assessment.loc[assessment["ID_CENARIO"] == "CBS_2026_BASE", "S_APUR"].iloc[0]) == Decimal("9.00")
    assert decimal_value(comparison.iloc[0]["DELTA_S_APUR"]) == Decimal("0.00")


def test_load_workbook_inputs_ignores_fiscal_derived_sheets(tmp_path):
    path, _ = build_case(tmp_path, cbs_workbook_inputs(cbs_tax_context(control=True)))
    wb = load_workbook(path)
    wb["FISCAL_RESULTADOS_OPERACAO"]["A2"] = "DERIVADO_ADULTERADO"
    wb["FISCAL_APURACAO"]["A2"] = "DERIVADO_ADULTERADO"
    wb["COMPARATIVO_CENARIOS"]["A2"] = "DERIVADO_ADULTERADO"
    wb.save(path)

    reloaded = load_workbook_inputs(path)

    assert reloaded.tax_context is not None
    assert list(reloaded.tax_context.tax_scenarios["ID_CENARIO"]) == ["CBS_2026_BASE", "CBS_2026_CONTROLE"]


def test_invalid_counterfactual_experiment_prevents_workbook_generation(tmp_path):
    output = tmp_path / "invalid_counterfactual.xlsx"

    with pytest.raises(SchemaValidationError):
        build_workbook(cbs_workbook_inputs(cbs_tax_context(control=True, invalid_control=True)), output)

    assert not output.exists()


def test_cash_recoding_in_workbook_inputs_flows_to_postings(tmp_path):
    chart_df = chart().copy()
    chart_df.loc[chart_df["COD_CTA"] == DEFAULT_ACCOUNT_ROLE_MAP["caixa"], "COD_CTA"] = "1.01.001.0001"
    account_mapping = build_default_account_role_mapping()
    account_mapping.loc[account_mapping["PAPEL_CONTABIL"] == "caixa", "COD_CTA"] = "1.01.001.0001"
    path, _ = build_case(tmp_path, workbook_inputs(chart_of_accounts=chart_df, account_role_mapping=account_mapping))
    postings = sheet_frame(path, "PARTIDAS")
    assert "1.01.001.0001" in set(postings["COD_CTA"])


def test_changing_statement_mapping_changes_presentation_not_postings(tmp_path):
    statement_mapping = build_default_statement_mapping(chart())
    statement_mapping.loc[statement_mapping["COD_CTA"] == "1.1.01.01", "COD_LINHA"] = "BP_BANCOS"
    path, _ = build_case(tmp_path, workbook_inputs(statement_mapping=statement_mapping))
    bp = sheet_frame(path, "BP")
    postings = sheet_frame(path, "PARTIDAS")
    assert decimal_value(row_by_line(bp, "BP_CAIXA")["VL"]) == Decimal("0.00")
    assert decimal_value(row_by_line(bp, "BP_BANCOS")["VL"]) == Decimal("100000.00")
    assert "1.1.01.01" in set(postings["COD_CTA"])


def test_changing_only_cod_df_is_ignored_and_overwritten_by_statement_mapping(tmp_path):
    chart_df = chart()
    chart_df.loc[chart_df["COD_CTA"] == "1.1.01.01", "COD_DF"] = "BP_BANCOS"
    statement_mapping = build_default_statement_mapping(chart())
    path, _ = build_case(tmp_path, workbook_inputs(chart_of_accounts=chart_df, statement_mapping=statement_mapping))
    bp = sheet_frame(path, "BP")
    chart_sheet = sheet_frame(path, "PLANO_CONTAS")
    assert decimal_value(row_by_line(bp, "BP_CAIXA")["VL"]) == Decimal("100000.00")
    assert decimal_value(row_by_line(bp, "BP_BANCOS")["VL"]) == Decimal("0.00")
    assert row_by_code(chart_sheet, "1.1.01.01")["COD_DF"] == "BP_CAIXA"


def test_same_input_generates_same_values_after_two_materializations(tmp_path):
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    inputs = workbook_inputs()
    build_workbook(inputs, first)
    build_workbook(inputs, second)
    for sheet_name in TABLE_NAMES:
        assert_frame_equal(sheet_frame(first, sheet_name), sheet_frame(second, sheet_name), check_dtype=False)


def test_build_workbook_does_not_mutate_input_dataframes(tmp_path):
    inputs = workbook_inputs()
    chart_copy = inputs.chart_of_accounts.copy(deep=True)
    mapping_copy = inputs.account_role_mapping.copy(deep=True)
    events_copy = inputs.events.copy(deep=True)
    build_workbook(inputs, tmp_path / "case.xlsx")
    assert_frame_equal(inputs.chart_of_accounts, chart_copy)
    assert_frame_equal(inputs.account_role_mapping, mapping_copy)
    assert_frame_equal(inputs.events, events_copy)


def test_workbook_does_not_need_formulas_to_close(tmp_path):
    path, _ = build_case(tmp_path)
    wb = load_workbook(path, data_only=False)
    formulas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append((ws.title, cell.coordinate, cell.value))
    assert formulas == []


def test_workbook_pipeline_matches_python_core_counts(tmp_path):
    inputs = workbook_inputs()
    path, _ = build_case(tmp_path, inputs)
    core = post_events(inputs.events, inputs.chart_of_accounts, inputs.simulation_config, account_role_mapping=inputs.account_role_mapping)
    assert len(sheet_frame(path, "LANCAMENTOS")) == len(core.journal_entry_headers)
    assert len(sheet_frame(path, "PARTIDAS")) == len(core.postings)
