"""Build and validate the Simples 2027 puro vs hibrido demo workbook."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
import sys
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from accounting_sim.account_mapping import build_default_account_role_mapping  # noqa: E402
from accounting_sim.canonical import (  # noqa: E402
    EVENT_COLUMNS,
    TAX_ANALYSIS_PARAMETER_COLUMNS,
    TAX_PARAMETER_COLUMNS,
    TAX_SCENARIO_COLUMNS,
    AccountingInvariantError,
    EventType,
    SimulationConfig,
)
from accounting_sim.chart_of_accounts import build_default_commercial_chart  # noqa: E402
from accounting_sim.statements import build_default_statement_mapping  # noqa: E402
from accounting_sim.tax_context import TaxContext  # noqa: E402
from accounting_sim.workbook import (  # noqa: E402
    TABLE_NAMES,
    WORKBOOK_SHEETS,
    WORKBOOK_SPEC_VERSION,
    WorkbookInputs,
    build_workbook,
)


ARTIFACT_PATH = PROJECT_ROOT / "artifacts" / "demo_simples_2027_puro_vs_hibrido.xlsx"
FIXTURE_DIR = PROJECT_ROOT / "data" / "examples" / "simples_2027"
BASE_SCENARIO_ID = "SIMPLES_2027_PURO"
ALT_SCENARIO_ID = "SIMPLES_2027_HIBRIDO"


@dataclass(frozen=True)
class Simples2027DemoSummary:
    path: Path
    size_bytes: int
    sheet_count: int
    row_counts: dict[str, int]
    baseline_charge: Decimal
    alternative_charge: Decimal
    delta_charge: Decimal


def build_simples_2027_demo(path: str | Path = ARTIFACT_PATH) -> Simples2027DemoSummary:
    output_path = Path(path)
    build_workbook(build_simples_2027_workbook_inputs(), output_path)
    return validate_simples_2027_demo(output_path)


def build_simples_2027_workbook_inputs() -> WorkbookInputs:
    period_start = date(2027, 1, 1)
    period_end = date(2027, 1, 31)
    chart = build_default_commercial_chart(period_start)
    config = SimulationConfig(
        simulation_id="DEMO_SIMPLES_2027",
        start_date=period_start,
        end_date=period_end,
        currency="BRL",
        seed=0,
        scenario_name="simples_2027_puro_vs_hibrido",
        spec_version=WORKBOOK_SPEC_VERSION,
    )
    return WorkbookInputs(
        simulation_config=config,
        chart_of_accounts=chart,
        account_role_mapping=build_default_account_role_mapping(),
        events=_load_events(),
        statement_mapping=build_default_statement_mapping(chart),
        tax_context=_load_tax_context(),
        tax_analysis_parameters=_load_analysis_parameters(),
    )


def validate_simples_2027_demo(path: str | Path) -> Simples2027DemoSummary:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True)
    if tuple(wb.sheetnames) != WORKBOOK_SHEETS:
        raise AccountingInvariantError("Workbook demo Simples 2027 fora da ordem de abas.")

    frames = {sheet_name: _read_table_frame(wb, sheet_name) for sheet_name in TABLE_NAMES}
    row_counts = {sheet_name: len(frame) for sheet_name, frame in frames.items()}
    for sheet_name, table_name in TABLE_NAMES.items():
        ws = wb[sheet_name]
        if row_counts[sheet_name] == 0:
            if table_name in ws.tables:
                raise AccountingInvariantError(f"Aba vazia {sheet_name} nao deve conter Table de uma linha.")
        elif table_name not in ws.tables:
            raise AccountingInvariantError(f"Tabela nomeada ausente em {sheet_name}.")
    validations = frames["VALIDACOES"]
    failed = validations.loc[validations["OK"] != True]  # noqa: E712
    if not failed.empty:
        raise AccountingInvariantError("VALIDACOES contem falhas no demo Simples 2027.")

    _assert_row_count(row_counts, "EVENTOS", 4)
    _assert_row_count(row_counts, "CENARIOS_TRIBUTARIOS", 2)
    _assert_row_count(row_counts, "ANALISE_PARAM", 2)
    _assert_row_count(row_counts, "SIMPLES_2027_RESULTADOS", 2)
    _assert_row_count(row_counts, "SIMPLES_2027_COMPARACAO", 5)
    _assert_row_count(row_counts, "FISCAL_RESULTADOS_OPERACAO", 0)
    _assert_row_count(row_counts, "FISCAL_APURACAO", 0)
    _assert_row_count(row_counts, "COMPARATIVO_CENARIOS", 0)

    results = frames["SIMPLES_2027_RESULTADOS"]
    comparison = frames["SIMPLES_2027_COMPARACAO"]
    puro = results.loc[results["ID_CENARIO"] == BASE_SCENARIO_ID].iloc[0]
    hibrido = results.loc[results["ID_CENARIO"] == ALT_SCENARIO_ID].iloc[0]
    encargo = comparison.loc[comparison["METRICA"] == "ENCARGO_TRIBUTARIO_COMPARAVEL"].iloc[0]

    if _money(puro["ENCARGO_TRIBUTARIO_COMPARAVEL"]) != Decimal("8825.00"):
        raise AccountingInvariantError("Encargo puro inesperado.")
    if _money(hibrido["ENCARGO_TRIBUTARIO_COMPARAVEL"]) != Decimal("8822.13"):
        raise AccountingInvariantError("Encargo hibrido inesperado.")
    if _money(encargo["DELTA"]) != Decimal("-2.87"):
        raise AccountingInvariantError("Delta de encargo inesperado.")
    if "CBS_2027_ANALYSIS_RATE_FRACTION" in set(frames["FISCAL_PARAM"]["CHAVE_PARAM"]):
        raise AccountingInvariantError("FISCAL_PARAM nao deve conter taxa CBS analitica.")
    _assert_no_formulas(workbook_path)
    _assert_no_vba(workbook_path)

    return Simples2027DemoSummary(
        path=workbook_path,
        size_bytes=workbook_path.stat().st_size,
        sheet_count=len(wb.sheetnames),
        row_counts=row_counts,
        baseline_charge=_money(puro["ENCARGO_TRIBUTARIO_COMPARAVEL"]),
        alternative_charge=_money(hibrido["ENCARGO_TRIBUTARIO_COMPARAVEL"]),
        delta_charge=_money(encargo["DELTA"]),
    )


def _load_tax_context() -> TaxContext:
    return TaxContext(
        entity_profile=pd.read_csv(FIXTURE_DIR / "entity_profile.csv", dtype=str, keep_default_na=False),
        fiscal_event_attributes=pd.read_csv(FIXTURE_DIR / "fiscal_event_attributes.csv", dtype=str, keep_default_na=False),
        tax_scenarios=_load_tax_scenarios(),
        tax_parameters=_load_tax_parameters(),
    )


def _load_events() -> pd.DataFrame:
    events = pd.read_csv(FIXTURE_DIR / "events.csv", dtype=str, keep_default_na=False)
    events["DT_EVENTO"] = pd.to_datetime(events["DT_EVENTO"]).dt.date
    events["VL_EVENTO_CENTS"] = events["VL_EVENTO_CENTS"].astype(int)
    events["VL_CUSTO_CENTS"] = events["VL_CUSTO_CENTS"].replace("", pd.NA)
    mask = events["VL_CUSTO_CENTS"].notna()
    events.loc[mask, "VL_CUSTO_CENTS"] = events.loc[mask, "VL_CUSTO_CENTS"].astype(int)
    for column in ("MEIO_FINANCEIRO", "CATEGORIA_DESPESA", "COD_PART", "DOC_REF"):
        events[column] = events[column].replace("", pd.NA)
    events.loc[events["TIPO_EVENTO"] == EventType.SALE_CREDIT.value, "MEIO_FINANCEIRO"] = pd.NA
    return events.loc[:, list(EVENT_COLUMNS)]


def _load_tax_scenarios() -> pd.DataFrame:
    scenarios = pd.read_csv(FIXTURE_DIR / "tax_scenarios.csv", dtype=str, keep_default_na=False)
    scenarios["DT_REFERENCIA_NORMATIVA"] = pd.to_datetime(scenarios["DT_REFERENCIA_NORMATIVA"]).dt.date
    scenarios["E_BASELINE"] = scenarios["E_BASELINE"].map(lambda value: str(value).lower() == "true")
    scenarios["ATIVO"] = scenarios["ATIVO"].map(lambda value: str(value).lower() == "true")
    return scenarios.loc[:, list(TAX_SCENARIO_COLUMNS)]


def _load_tax_parameters() -> pd.DataFrame:
    parameters = pd.read_csv(FIXTURE_DIR / "tax_parameters.csv", dtype=str, keep_default_na=False)
    for column in ("VIG_INI", "DATA_CONSULTA"):
        parameters[column] = pd.to_datetime(parameters[column]).dt.date
    parameters["VIG_FIM"] = parameters["VIG_FIM"].map(
        lambda value: None if str(value).strip() == "" else pd.to_datetime(value).date()
    )
    return parameters.loc[:, list(TAX_PARAMETER_COLUMNS)]


def _load_analysis_parameters() -> pd.DataFrame:
    return pd.read_csv(FIXTURE_DIR / "analysis_parameters.csv", dtype=str, keep_default_na=False).loc[
        :, list(TAX_ANALYSIS_PARAMETER_COLUMNS)
    ]


def _read_table_frame(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    if TABLE_NAMES[sheet_name] in ws.tables:
        table = ws.tables[TABLE_NAMES[sheet_name]]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    else:
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0], dtype=object)


def _assert_row_count(row_counts: dict[str, int], sheet_name: str, expected: int) -> None:
    if row_counts[sheet_name] != expected:
        raise AccountingInvariantError(f"{sheet_name} deveria ter {expected} linhas.")


def _assert_no_formulas(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    raise AccountingInvariantError(f"Formula encontrada em {ws.title}!{cell.coordinate}.")


def _assert_no_vba(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        if any(name.endswith("vbaProject.bin") for name in archive.namelist()):
            raise AccountingInvariantError("Arquivo xlsx nao deve conter VBA.")


def _money(value: object) -> Decimal:
    if value is None:
        raise AccountingInvariantError("Valor monetario esperado veio vazio.")
    return Decimal(str(value)).quantize(Decimal("0.00"))


def main() -> None:
    summary = build_simples_2027_demo()
    print(f"generated={summary.path}")
    print(f"size_bytes={summary.size_bytes}")
    print(f"sheet_count={summary.sheet_count}")
    print(f"baseline.ENCARGO={summary.baseline_charge}")
    print(f"alternative.ENCARGO={summary.alternative_charge}")
    print(f"comparison.DELTA_ENCARGO={summary.delta_charge}")


if __name__ == "__main__":
    main()
