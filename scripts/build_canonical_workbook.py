"""Build and validate the canonical end-to-end workbook artifact."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
import sys
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from accounting_sim.account_mapping import build_default_account_role_mapping  # noqa: E402
from accounting_sim.canonical import (  # noqa: E402
    EVENT_COLUMNS,
    FISCAL_EVENT_ATTRIBUTE_COLUMNS,
    TAX_PARAMETER_COLUMNS,
    TAX_SCENARIO_COLUMNS,
    AccountingInvariantError,
    EventType,
    SimulationConfig,
)
from accounting_sim.chart_of_accounts import build_default_commercial_chart  # noqa: E402
from accounting_sim.statements import build_default_statement_mapping  # noqa: E402
from accounting_sim.tax_context import TaxContext  # noqa: E402
from accounting_sim.workbook import (  # noqa: E402
    EDITABLE_SHEETS,
    TABLE_NAMES,
    WORKBOOK_SHEETS,
    WORKBOOK_SPEC_VERSION,
    WorkbookInputs,
    build_workbook,
)


ARTIFACT_PATH = PROJECT_ROOT / "artifacts" / "contabilidade_parametrizada.xlsx"
CBS_FIXTURE_DIR = PROJECT_ROOT / "data" / "examples" / "cbs_2026"
BASE_SCENARIO_ID = "CBS_2026_BASE"
CONTROL_SCENARIO_ID = "CBS_2026_CONTROLE"


@dataclass(frozen=True)
class WorkbookArtifactSummary:
    path: Path
    size_bytes: int
    sheet_count: int
    row_counts: dict[str, int]
    baseline_assessment: dict[str, object]
    control_comparison: dict[str, object]


def build_canonical_workbook(path: str | Path = ARTIFACT_PATH) -> WorkbookArtifactSummary:
    output_path = Path(path)
    inputs = build_canonical_workbook_inputs()
    build_workbook(inputs, output_path)
    return validate_canonical_workbook(output_path)


def build_canonical_workbook_inputs() -> WorkbookInputs:
    period_start = date(2026, 8, 1)
    period_end = date(2026, 8, 31)
    chart = build_default_commercial_chart(period_start)
    config = SimulationConfig(
        simulation_id="CANONICAL_CBS_2026",
        start_date=period_start,
        end_date=period_end,
        currency="BRL",
        seed=0,
        scenario_name="cbs_2026_counterfactual_control",
        spec_version=WORKBOOK_SPEC_VERSION,
    )
    return WorkbookInputs(
        simulation_config=config,
        chart_of_accounts=chart,
        account_role_mapping=build_default_account_role_mapping(),
        events=_load_events(),
        statement_mapping=build_default_statement_mapping(chart),
        tax_context=_load_tax_context(),
    )


def validate_canonical_workbook(path: str | Path) -> WorkbookArtifactSummary:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True)
    if tuple(wb.sheetnames) != WORKBOOK_SHEETS:
        raise AccountingInvariantError("Workbook canonico fora da ordem de abas.")
    for sheet_name, table_name in TABLE_NAMES.items():
        ws = wb[sheet_name]
        if table_name not in ws.tables:
            raise AccountingInvariantError(f"Tabela nomeada ausente em {sheet_name}.")
        if ws.freeze_panes != "A2":
            raise AccountingInvariantError(f"Freeze panes invalido em {sheet_name}.")
        if ws.tables[table_name].autoFilter is None:
            raise AccountingInvariantError(f"Filtro ausente em {sheet_name}.")

    if not EDITABLE_SHEETS.issubset(set(WORKBOOK_SHEETS)):
        raise AccountingInvariantError("Abas editaveis fora do contrato canonico.")
    if {"FISCAL_RESULTADOS_OPERACAO", "FISCAL_APURACAO", "COMPARATIVO_CENARIOS", "SIMPLES_2027_RESULTADOS", "SIMPLES_2027_COMPARACAO"} & set(EDITABLE_SHEETS):
        raise AccountingInvariantError("Abas fiscais derivadas nao podem ser editaveis.")

    frames = {sheet_name: _read_table_frame(wb, sheet_name) for sheet_name in TABLE_NAMES}
    row_counts = {sheet_name: len(frame) for sheet_name, frame in frames.items()}

    validations = frames["VALIDACOES"]
    failed = validations.loc[validations["OK"] != True]  # noqa: E712
    if not failed.empty:
        raise AccountingInvariantError("VALIDACOES contem falhas.")

    postings = frames["PARTIDAS"]
    debit_total = sum(_money(value) for value in postings.loc[postings["IND_DC"] == "D", "VL_DC"])
    credit_total = sum(_money(value) for value in postings.loc[postings["IND_DC"] == "C", "VL_DC"])
    if debit_total != credit_total:
        raise AccountingInvariantError("Partidas nao fecham por debito e credito.")

    if row_counts["LANCAMENTOS"] <= 0 or row_counts["BALANCETE"] <= 0 or row_counts["BP"] <= 0 or row_counts["DRE"] <= 0:
        raise AccountingInvariantError("Nucleo contabil ou demonstracoes ausentes.")

    _assert_row_count(row_counts, "EVENTOS", 3)
    _assert_row_count(row_counts, "FISCAL_RESULTADOS_OPERACAO", 4)
    _assert_row_count(row_counts, "FISCAL_APURACAO", 2)
    _assert_row_count(row_counts, "COMPARATIVO_CENARIOS", 1)
    _assert_row_count(row_counts, "ANALISE_PARAM", 0)
    _assert_row_count(row_counts, "SIMPLES_2027_RESULTADOS", 0)
    _assert_row_count(row_counts, "SIMPLES_2027_COMPARACAO", 0)

    scenarios = frames["CENARIOS_TRIBUTARIOS"]
    if set(scenarios["ID_CENARIO"]) != {BASE_SCENARIO_ID, CONTROL_SCENARIO_ID}:
        raise AccountingInvariantError("Cenarios fiscais canonicos ausentes.")

    fiscal_operations = frames["FISCAL_RESULTADOS_OPERACAO"]
    _assert_number_format(wb, "FISCAL_RESULTADOS_OPERACAO", "ALIQUOTA", "0.0000%")
    for column in ("BASE", "CREDITO", "DEBITO"):
        _assert_number_format(wb, "FISCAL_RESULTADOS_OPERACAO", column, "#,##0.00")
    for column in ("S_APUR", "T_RECOLHER", "P_CASH", "E_DRE", "C_SALDO"):
        _assert_number_format(wb, "FISCAL_APURACAO", column, "#,##0.00")
    for column in ("DELTA_S_APUR", "DELTA_T_RECOLHER", "DELTA_P_CASH", "DELTA_E_DRE", "DELTA_C_SALDO"):
        _assert_number_format(wb, "COMPARATIVO_CENARIOS", column, "#,##0.00")
    if Decimal(str(fiscal_operations.iloc[0]["ALIQUOTA"])) != Decimal("0.009"):
        raise AccountingInvariantError("Aliquota CBS deve permanecer como fracao no workbook.")

    assessment = frames["FISCAL_APURACAO"]
    baseline = assessment.loc[assessment["ID_CENARIO"] == BASE_SCENARIO_ID].iloc[0].to_dict()
    for scenario_id in (BASE_SCENARIO_ID, CONTROL_SCENARIO_ID):
        row = assessment.loc[assessment["ID_CENARIO"] == scenario_id].iloc[0]
        if _money(row["S_APUR"]) != Decimal("9.00"):
            raise AccountingInvariantError("S_APUR canonico inesperado.")
        if _money(row["T_RECOLHER"]) != Decimal("0.00"):
            raise AccountingInvariantError("T_RECOLHER canonico inesperado.")
        if row["P_CASH"] is not None or row["E_DRE"] is not None:
            raise AccountingInvariantError("P_CASH e E_DRE devem permanecer vazios.")
        if _money(row["C_SALDO"]) != Decimal("0.00"):
            raise AccountingInvariantError("C_SALDO canonico inesperado.")

    comparison = frames["COMPARATIVO_CENARIOS"].iloc[0].to_dict()
    if comparison["ID_CENARIO_BASE"] != BASE_SCENARIO_ID or comparison["ID_CENARIO"] != CONTROL_SCENARIO_ID:
        raise AccountingInvariantError("Comparativo canonico deve ser controle contra baseline.")
    if _money(comparison["DELTA_S_APUR"]) != Decimal("0.00"):
        raise AccountingInvariantError("DELTA_S_APUR canonico inesperado.")
    if _money(comparison["DELTA_T_RECOLHER"]) != Decimal("0.00"):
        raise AccountingInvariantError("DELTA_T_RECOLHER canonico inesperado.")
    if comparison["DELTA_P_CASH"] is not None or comparison["DELTA_E_DRE"] is not None:
        raise AccountingInvariantError("Deltas desconhecidos devem permanecer vazios.")
    if _money(comparison["DELTA_C_SALDO"]) != Decimal("0.00"):
        raise AccountingInvariantError("DELTA_C_SALDO canonico inesperado.")

    _assert_no_formulas(workbook_path)
    _assert_no_vba(workbook_path)

    return WorkbookArtifactSummary(
        path=workbook_path,
        size_bytes=workbook_path.stat().st_size,
        sheet_count=len(wb.sheetnames),
        row_counts=row_counts,
        baseline_assessment=baseline,
        control_comparison=comparison,
    )


def _load_tax_context() -> TaxContext:
    return TaxContext(
        entity_profile=pd.read_csv(CBS_FIXTURE_DIR / "entity_profile.csv", dtype=str, keep_default_na=False),
        fiscal_event_attributes=pd.read_csv(CBS_FIXTURE_DIR / "fiscal_event_attributes.csv", dtype=str, keep_default_na=False),
        tax_scenarios=_load_tax_scenarios_with_control(),
        tax_parameters=_load_tax_parameters(),
    )


def _load_events() -> pd.DataFrame:
    events = pd.read_csv(CBS_FIXTURE_DIR / "events.csv", dtype=str, keep_default_na=False)
    events["DT_EVENTO"] = pd.to_datetime(events["DT_EVENTO"]).dt.date
    events["VL_EVENTO_CENTS"] = events["VL_EVENTO_CENTS"].astype(int)
    events["VL_CUSTO_CENTS"] = events["VL_CUSTO_CENTS"].replace("", pd.NA)
    mask = events["VL_CUSTO_CENTS"].notna()
    events.loc[mask, "VL_CUSTO_CENTS"] = events.loc[mask, "VL_CUSTO_CENTS"].astype(int)
    for column in ("MEIO_FINANCEIRO", "CATEGORIA_DESPESA", "COD_PART", "DOC_REF"):
        events[column] = events[column].replace("", pd.NA)
    events.loc[events["TIPO_EVENTO"] == EventType.SALE_CREDIT.value, "MEIO_FINANCEIRO"] = pd.NA
    return events.loc[:, list(EVENT_COLUMNS)]


def _load_tax_scenarios_with_control() -> pd.DataFrame:
    scenarios = pd.read_csv(CBS_FIXTURE_DIR / "tax_scenarios.csv", dtype=str, keep_default_na=False)
    scenarios["DT_REFERENCIA_NORMATIVA"] = pd.to_datetime(scenarios["DT_REFERENCIA_NORMATIVA"]).dt.date
    scenarios["E_BASELINE"] = scenarios["E_BASELINE"].map(lambda value: str(value).lower() == "true")
    scenarios["ATIVO"] = scenarios["ATIVO"].map(lambda value: str(value).lower() == "true")
    if CONTROL_SCENARIO_ID not in set(scenarios["ID_CENARIO"]):
        control = scenarios.loc[scenarios["ID_CENARIO"] == BASE_SCENARIO_ID].iloc[0].copy()
        control["ID_CENARIO"] = CONTROL_SCENARIO_ID
        control["DESCRICAO"] = "controle estrutural de orquestracao"
        control["E_BASELINE"] = False
        control["ATIVO"] = True
        scenarios = pd.concat([scenarios, pd.DataFrame([control])], ignore_index=True)
    return scenarios.loc[:, list(TAX_SCENARIO_COLUMNS)]


def _load_tax_parameters() -> pd.DataFrame:
    parameters = pd.read_csv(CBS_FIXTURE_DIR / "tax_parameters.csv", dtype=str, keep_default_na=False)
    for column in ("VIG_INI", "DATA_CONSULTA"):
        parameters[column] = pd.to_datetime(parameters[column]).dt.date
    parameters["VIG_FIM"] = parameters["VIG_FIM"].map(
        lambda value: None if str(value).strip() == "" else pd.to_datetime(value).date()
    )
    return parameters.loc[:, list(TAX_PARAMETER_COLUMNS)]


def _read_table_frame(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    table = ws.tables[TABLE_NAMES[sheet_name]]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0], dtype=object)


def _assert_row_count(row_counts: dict[str, int], sheet_name: str, expected: int) -> None:
    if row_counts[sheet_name] != expected:
        raise AccountingInvariantError(f"{sheet_name} deveria ter {expected} linhas.")


def _assert_number_format(wb, sheet_name: str, column: str, expected: str) -> None:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    column_index = headers.index(column) + 1
    if ws.cell(row=2, column=column_index).number_format != expected:
        raise AccountingInvariantError(f"Formato invalido em {sheet_name}.{column}.")


def _assert_no_formulas(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    raise AccountingInvariantError(f"Formula encontrada em {ws.title}!{cell.coordinate}.")


def _assert_no_vba(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        if any(name.endswith("vbaProject.bin") for name in archive.namelist()):
            raise AccountingInvariantError("Arquivo xlsx nao deve conter VBA.")


def _money(value: object) -> Decimal:
    if value is None:
        raise AccountingInvariantError("Valor monetario esperado veio vazio.")
    return Decimal(str(value)).quantize(Decimal("0.00"))


def main() -> None:
    summary = build_canonical_workbook()
    print(f"generated={summary.path}")
    print(f"size_bytes={summary.size_bytes}")
    print(f"sheet_count={summary.sheet_count}")
    for sheet_name in (
        "EVENTOS",
        "LANCAMENTOS",
        "PARTIDAS",
        "BALANCETE",
        "BP",
        "DRE",
        "FISCAL_RESULTADOS_OPERACAO",
        "FISCAL_APURACAO",
        "COMPARATIVO_CENARIOS",
    ):
        print(f"rows.{sheet_name}={summary.row_counts[sheet_name]}")
    print(f"baseline.S_APUR={summary.baseline_assessment['S_APUR']}")
    print(f"baseline.T_RECOLHER={summary.baseline_assessment['T_RECOLHER']}")
    print(f"comparison.DELTA_S_APUR={summary.control_comparison['DELTA_S_APUR']}")
    print(f"comparison.DELTA_T_RECOLHER={summary.control_comparison['DELTA_T_RECOLHER']}")


if __name__ == "__main__":
    main()
