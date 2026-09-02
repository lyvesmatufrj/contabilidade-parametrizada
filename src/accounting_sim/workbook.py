"""Materialização Excel do núcleo contábil e interface tributária da spec 08."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from accounting_sim.account_mapping import validate_account_role_mapping
from accounting_sim.canonical import (
    ACCOUNT_ROLE_MAPPING_COLUMNS,
    BALANCE_SHEET_COLUMNS,
    CHART_OF_ACCOUNTS_COLUMNS,
    ENTITY_PROFILE_COLUMNS,
    EVENT_COLUMNS,
    EVENT_ENTRY_LINK_COLUMNS,
    FISCAL_EVENT_ATTRIBUTE_COLUMNS,
    INCOME_STATEMENT_COLUMNS,
    JOURNAL_ENTRY_HEADER_COLUMNS,
    JOURNAL_VIEW_COLUMNS,
    LEDGER_VIEW_COLUMNS,
    POSTING_COLUMNS,
    TAX_PARAMETER_COLUMNS,
    TAX_SCENARIO_COLUMNS,
    STATEMENT_MAPPING_COLUMNS,
    TRIAL_BALANCE_COLUMNS,
    AccountingInvariantError,
    AccountingPeriod,
    DebitCredit,
    EventClass,
    EventDirection,
    EventNature,
    EventType,
    Origin,
    PaymentTerm,
    SchemaValidationError,
    SimulationConfig,
    ValidationIssue,
    ValidationReport,
    parse_iso_date,
)
from accounting_sim.chart_of_accounts import validate_chart_of_accounts
from accounting_sim.events import normalize_events, sort_events, validate_events
from accounting_sim.ledger import (
    build_journal,
    build_ledger,
    build_trial_balance,
    validate_ledger_trial_balance,
)
from accounting_sim.posting import post_events, validate_posting_result
from accounting_sim.statements import (
    FINANCIAL_STATEMENT_SPEC_VERSION,
    build_default_statement_mapping,
    build_financial_statements,
    synchronize_chart_statement_codes,
    validate_financial_statements,
    validate_statement_mapping,
)
from accounting_sim.tax_context import (
    TAX_INTERFACE_SPEC_VERSION,
    TaxContext,
    build_empty_tax_context,
    validate_entity_profile,
    validate_fiscal_event_attributes,
    validate_tax_context,
    validate_tax_parameters,
    validate_tax_scenarios,
)


WORKBOOK_SPEC_VERSION = "spec_08_excel_workbook_v1"

WORKBOOK_SHEETS: tuple[str, ...] = (
    "README",
    "CONFIG",
    "ENTIDADE",
    "PLANO_CONTAS",
    "MAPEAMENTO_CONTAS",
    "EVENTOS",
    "EVENTOS_FISCAIS",
    "LANCAMENTOS",
    "PARTIDAS",
    "VINCULO_EVENTO_LCTO",
    "DIARIO",
    "RAZAO",
    "BALANCETE",
    "MAPEAMENTO_DF",
    "BP",
    "DRE",
    "CENARIOS_TRIBUTARIOS",
    "FISCAL_PARAM",
    "VALIDACOES",
    "PROVENIENCIA",
)

EDITABLE_SHEETS = frozenset(
    {
        "CONFIG",
        "ENTIDADE",
        "PLANO_CONTAS",
        "MAPEAMENTO_CONTAS",
        "EVENTOS",
        "EVENTOS_FISCAIS",
        "MAPEAMENTO_DF",
        "CENARIOS_TRIBUTARIOS",
        "FISCAL_PARAM",
    }
)

CONFIG_COLUMNS: tuple[str, ...] = ("CHAVE", "VALOR")
VALIDATION_COLUMNS: tuple[str, ...] = (
    "ETAPA",
    "OK",
    "ISSUE_CODE",
    "MENSAGEM",
    "ACCOUNT_CODE",
    "EVENT_ID",
    "ENTRY_ID",
    "POSTING_ID",
    "ENTITY_ID",
    "SCENARIO_ID",
    "TAX_PARAM_ID",
)
PROVENANCE_COLUMNS: tuple[str, ...] = ("CHAVE", "VALOR")

EVENT_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "ID_EVENTO",
    "DT_EVENTO",
    "CLASSE_EVENTO",
    "TIPO_EVENTO",
    "DIRECAO",
    "NATUREZA",
    "VL_EVENTO",
    "VL_CUSTO",
    "MEIO_FINANCEIRO",
    "CATEGORIA_DESPESA",
    "COD_PART",
    "COND_PAGTO",
    "DOC_REF",
    "HIST",
    "ORIGEM",
    "SPEC_VERSION",
)
JOURNAL_ENTRY_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "NUM_LCTO",
    "DT_LCTO",
    "VL_LCTO",
    "IND_LCTO",
    "DT_LCTO_EXT",
    "ID_GERACAO",
    "VERSAO_REGRA",
)
POSTING_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "ID_PARTIDA",
    "NUM_LCTO",
    "COD_CTA",
    "COD_CCUS",
    "VL_DC",
    "IND_DC",
    "NUM_ARQ",
    "COD_HIST_PAD",
    "HIST",
    "COD_PART",
    "ID_ORIGEM",
)
JOURNAL_VIEW_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "DT_LCTO",
    "NUM_LCTO",
    "ID_PARTIDA",
    "COD_CTA",
    "CTA",
    "IND_DC",
    "VL_DC",
    "HIST",
    "COD_PART",
    "ID_ORIGEM",
)
LEDGER_VIEW_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "COD_CTA",
    "CTA",
    "DT_LCTO",
    "NUM_LCTO",
    "ID_PARTIDA",
    "DEBITO",
    "CREDITO",
    "MOVIMENTO_ASSINADO",
    "SALDO_ASSINADO",
    "SALDO_ABS",
    "IND_DC_SALDO",
    "HIST",
    "ID_ORIGEM",
)
TRIAL_BALANCE_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "DT_INI",
    "DT_FIN",
    "COD_CTA",
    "COD_CCUS",
    "VL_SLD_INI",
    "IND_DC_INI",
    "VL_DEB",
    "VL_CRED",
    "VL_SLD_FIN",
    "IND_DC_FIN",
)
BALANCE_SHEET_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "DT_REF",
    "ORDEM",
    "COD_LINHA",
    "NIVEL",
    "TIPO_LINHA",
    "LINHA",
    "VL",
)
INCOME_STATEMENT_WORKBOOK_COLUMNS: tuple[str, ...] = (
    "DT_INI",
    "DT_FIN",
    "ORDEM",
    "COD_LINHA",
    "NIVEL",
    "TIPO_LINHA",
    "LINHA",
    "VL",
)

TABLE_NAMES: Mapping[str, str] = {
    "CONFIG": "tbl_CONFIG",
    "ENTIDADE": "tbl_ENTIDADE",
    "PLANO_CONTAS": "tbl_PLANO_CONTAS",
    "MAPEAMENTO_CONTAS": "tbl_MAPEAMENTO_CONTAS",
    "EVENTOS": "tbl_EVENTOS",
    "EVENTOS_FISCAIS": "tbl_EVENTOS_FISCAIS",
    "LANCAMENTOS": "tbl_LANCAMENTOS",
    "PARTIDAS": "tbl_PARTIDAS",
    "VINCULO_EVENTO_LCTO": "tbl_VINCULO_EVENTO_LCTO",
    "DIARIO": "tbl_DIARIO",
    "RAZAO": "tbl_RAZAO",
    "BALANCETE": "tbl_BALANCETE",
    "MAPEAMENTO_DF": "tbl_MAPEAMENTO_DF",
    "BP": "tbl_BP",
    "DRE": "tbl_DRE",
    "CENARIOS_TRIBUTARIOS": "tbl_CENARIOS_TRIBUTARIOS",
    "FISCAL_PARAM": "tbl_FISCAL_PARAM",
    "VALIDACOES": "tbl_VALIDACOES",
    "PROVENIENCIA": "tbl_PROVENIENCIA",
}


@dataclass(frozen=True)
class WorkbookInputs:
    simulation_config: SimulationConfig
    chart_of_accounts: pd.DataFrame
    account_role_mapping: pd.DataFrame
    events: pd.DataFrame
    statement_mapping: pd.DataFrame | None = None
    tax_context: TaxContext | None = None


def build_workbook(
    inputs: WorkbookInputs,
    path: str | Path,
    *,
    rule_version: str = "posting_rules_v1",
) -> Path:
    simulation_config = inputs.simulation_config
    period = AccountingPeriod(simulation_config.start_date, simulation_config.end_date)
    tax_context = _copy_tax_context(inputs.tax_context) if inputs.tax_context is not None else build_empty_tax_context()
    raw_chart_of_accounts = inputs.chart_of_accounts.copy(deep=True)
    account_role_mapping = inputs.account_role_mapping.copy(deep=True)
    statement_mapping = (
        build_default_statement_mapping(raw_chart_of_accounts)
        if inputs.statement_mapping is None
        else inputs.statement_mapping.copy(deep=True)
    )
    chart_of_accounts = synchronize_chart_statement_codes(raw_chart_of_accounts, statement_mapping)
    events = sort_events(normalize_events(inputs.events.copy(deep=True)))

    chart_report = validate_chart_of_accounts(chart_of_accounts)
    mapping_report = validate_account_role_mapping(account_role_mapping, chart_of_accounts)
    statement_mapping_report = validate_statement_mapping(statement_mapping, chart_of_accounts)
    event_report = validate_events(events, period)
    entity_report = validate_entity_profile(tax_context.entity_profile)
    fiscal_event_report = validate_fiscal_event_attributes(tax_context.fiscal_event_attributes, events)
    tax_parameter_report = validate_tax_parameters(tax_context.tax_parameters)
    tax_scenario_report = validate_tax_scenarios(tax_context.tax_scenarios, tax_context.entity_profile, tax_context.tax_parameters)
    tax_context_report = validate_tax_context(tax_context, events)
    _raise_if_invalid("PLANO_CONTAS", chart_report)
    _raise_if_invalid("MAPEAMENTO_CONTAS", mapping_report)
    _raise_if_invalid("MAPEAMENTO_DF", statement_mapping_report)
    _raise_if_invalid("EVENTOS", event_report)
    _raise_if_invalid("ENTIDADE", entity_report)
    _raise_if_invalid("EVENTOS_FISCAIS", fiscal_event_report)
    _raise_if_invalid("FISCAL_PARAM", tax_parameter_report)
    _raise_if_invalid("CENARIOS_TRIBUTARIOS", tax_scenario_report)
    _raise_if_invalid("CONTEXTO_TRIBUTARIO", tax_context_report)

    posting_result = post_events(
        events,
        chart_of_accounts,
        simulation_config,
        account_role_mapping=account_role_mapping,
        rule_version=rule_version,
    )
    posting_report = validate_posting_result(posting_result, events, chart_of_accounts)
    _raise_if_invalid("LANCAMENTOS_PARTIDAS", posting_report)

    journal = build_journal(posting_result.journal_entry_headers, posting_result.postings, chart_of_accounts)
    ledger = build_ledger(posting_result.journal_entry_headers, posting_result.postings, chart_of_accounts)
    trial_balance = build_trial_balance(ledger, chart_of_accounts, period)
    ledger_report = validate_ledger_trial_balance(posting_result.postings, ledger, trial_balance, chart_of_accounts, period)
    _raise_if_invalid("RAZAO_BALANCETE", ledger_report)
    financial_statements = build_financial_statements(trial_balance, chart_of_accounts, statement_mapping, period)
    statements_report = validate_financial_statements(financial_statements, trial_balance, chart_of_accounts, statement_mapping, period)
    _raise_if_invalid("DEMONSTRACOES", statements_report)

    validations = _build_validations(
        {
            "PLANO_CONTAS": chart_report,
            "MAPEAMENTO_CONTAS": mapping_report,
            "MAPEAMENTO_DF": statement_mapping_report,
            "EVENTOS": event_report,
            "ENTIDADE": entity_report,
            "EVENTOS_FISCAIS": fiscal_event_report,
            "CENARIOS_TRIBUTARIOS": tax_scenario_report,
            "FISCAL_PARAM": tax_parameter_report,
            "CONTEXTO_TRIBUTARIO": tax_context_report,
            "LANCAMENTOS_PARTIDAS": posting_report,
            "RAZAO_BALANCETE": ledger_report,
            "DEMONSTRACOES": statements_report,
        }
    )
    provenance = _build_provenance(simulation_config, events, rule_version, tax_context)

    wb = Workbook()
    wb.remove(wb.active)
    _write_readme(wb)
    _write_table(wb, "CONFIG", _config_to_frame(simulation_config))
    _write_table(wb, "ENTIDADE", _serialize_frame(tax_context.entity_profile, ENTITY_PROFILE_COLUMNS))
    _write_table(wb, "PLANO_CONTAS", _serialize_frame(chart_of_accounts, CHART_OF_ACCOUNTS_COLUMNS))
    _write_table(wb, "MAPEAMENTO_CONTAS", _serialize_frame(account_role_mapping, ACCOUNT_ROLE_MAPPING_COLUMNS))
    _write_table(wb, "EVENTOS", _events_to_workbook(events))
    _write_table(wb, "EVENTOS_FISCAIS", _serialize_frame(tax_context.fiscal_event_attributes, FISCAL_EVENT_ATTRIBUTE_COLUMNS))
    _write_table(wb, "LANCAMENTOS", _journal_entries_to_workbook(posting_result.journal_entry_headers))
    _write_table(wb, "PARTIDAS", _postings_to_workbook(posting_result.postings))
    _write_table(wb, "VINCULO_EVENTO_LCTO", _serialize_frame(posting_result.event_entry_links, EVENT_ENTRY_LINK_COLUMNS))
    _write_table(wb, "DIARIO", _journal_to_workbook(journal))
    _write_table(wb, "RAZAO", _ledger_to_workbook(ledger))
    _write_table(wb, "BALANCETE", _trial_balance_to_workbook(trial_balance))
    _write_table(wb, "MAPEAMENTO_DF", _serialize_frame(statement_mapping, STATEMENT_MAPPING_COLUMNS))
    _write_table(wb, "BP", _balance_sheet_to_workbook(financial_statements.balance_sheet))
    _write_table(wb, "DRE", _income_statement_to_workbook(financial_statements.income_statement))
    _write_table(wb, "CENARIOS_TRIBUTARIOS", _serialize_frame(tax_context.tax_scenarios, TAX_SCENARIO_COLUMNS))
    _write_table(wb, "FISCAL_PARAM", _serialize_frame(tax_context.tax_parameters, TAX_PARAMETER_COLUMNS))
    _write_table(wb, "VALIDACOES", validations)
    _write_table(wb, "PROVENIENCIA", provenance)
    _apply_editable_validations(wb)

    if tuple(wb.sheetnames) != WORKBOOK_SHEETS:
        raise AccountingInvariantError("Workbook gerado com abas fora da ordem canônica.")

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def load_workbook_inputs(path: str | Path) -> WorkbookInputs:
    wb = load_workbook(path, data_only=True)
    for sheet_name in EDITABLE_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise SchemaValidationError(f"Aba de entrada ausente: {sheet_name}.")

    config = _frame_to_config(_read_table_frame(wb["CONFIG"], CONFIG_COLUMNS))
    entity_profile = _frame_to_entity_profile(_read_table_frame(wb["ENTIDADE"], ENTITY_PROFILE_COLUMNS))
    chart_of_accounts = _frame_to_chart(_read_table_frame(wb["PLANO_CONTAS"], CHART_OF_ACCOUNTS_COLUMNS))
    account_role_mapping = _read_table_frame(wb["MAPEAMENTO_CONTAS"], ACCOUNT_ROLE_MAPPING_COLUMNS)
    events = _frame_to_events(_read_table_frame(wb["EVENTOS"], EVENT_WORKBOOK_COLUMNS))
    fiscal_event_attributes = _frame_to_fiscal_event_attributes(_read_table_frame(wb["EVENTOS_FISCAIS"], FISCAL_EVENT_ATTRIBUTE_COLUMNS))
    statement_mapping = _read_table_frame(wb["MAPEAMENTO_DF"], STATEMENT_MAPPING_COLUMNS)
    tax_scenarios = _frame_to_tax_scenarios(_read_table_frame(wb["CENARIOS_TRIBUTARIOS"], TAX_SCENARIO_COLUMNS))
    tax_parameters = _frame_to_tax_parameters(_read_table_frame(wb["FISCAL_PARAM"], TAX_PARAMETER_COLUMNS))
    tax_context = TaxContext(
        entity_profile=entity_profile,
        fiscal_event_attributes=fiscal_event_attributes,
        tax_scenarios=tax_scenarios,
        tax_parameters=tax_parameters,
    )
    return WorkbookInputs(config, chart_of_accounts, account_role_mapping, events, statement_mapping, tax_context)


def regenerate_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    rule_version: str = "posting_rules_v1",
) -> Path:
    source_path = Path(input_path)
    inputs = load_workbook_inputs(source_path)
    destination = source_path if output_path is None else Path(output_path)
    return build_workbook(inputs, destination, rule_version=rule_version)


def _config_to_frame(config: SimulationConfig) -> pd.DataFrame:
    rows = [
        ("simulation_id", config.simulation_id),
        ("start_date", config.start_date),
        ("end_date", config.end_date),
        ("currency", config.currency),
        ("seed", config.seed),
        ("scenario_name", config.scenario_name),
        ("spec_version", config.spec_version),
    ]
    return pd.DataFrame(rows, columns=CONFIG_COLUMNS, dtype=object)


def _frame_to_config(frame: pd.DataFrame) -> SimulationConfig:
    values = {str(row["CHAVE"]).strip(): row["VALOR"] for _, row in frame.iterrows()}
    required = {"simulation_id", "start_date", "end_date", "currency", "seed", "scenario_name", "spec_version"}
    missing = required - set(values)
    if missing:
        raise SchemaValidationError(f"CONFIG sem chaves obrigatórias: {sorted(missing)}.")
    return SimulationConfig(
        simulation_id=str(values["simulation_id"]).strip(),
        start_date=_coerce_excel_date(values["start_date"]),
        end_date=_coerce_excel_date(values["end_date"]),
        currency=str(values["currency"]).strip(),
        seed=_coerce_int(values["seed"], "seed"),
        scenario_name=str(values["scenario_name"]).strip(),
        spec_version=str(values["spec_version"]).strip(),
    )


def _frame_to_chart(frame: pd.DataFrame) -> pd.DataFrame:
    chart = frame.copy()
    chart["DT_ALT"] = chart["DT_ALT"].map(_coerce_excel_date)
    chart["NIVEL"] = chart["NIVEL"].map(lambda value: _coerce_int(value, "NIVEL"))
    chart["ATIVA"] = chart["ATIVA"].map(_coerce_bool)
    for column in ("COD_CTA_SUP", "COD_DF"):
        chart[column] = chart[column].map(_blank_to_none)
    return chart.loc[:, list(CHART_OF_ACCOUNTS_COLUMNS)]


def _frame_to_events(frame: pd.DataFrame) -> pd.DataFrame:
    events = frame.copy()
    events["DT_EVENTO"] = events["DT_EVENTO"].map(_coerce_excel_date)
    events["VL_EVENTO_CENTS"] = pd.Series(
        [_excel_money_to_cents(value, optional=False) for value in events["VL_EVENTO"]],
        index=events.index,
        dtype=object,
    )
    events["VL_CUSTO_CENTS"] = pd.Series(
        [_excel_money_to_cents(value, optional=True) for value in events["VL_CUSTO"]],
        index=events.index,
        dtype=object,
    )
    events = events.drop(columns=["VL_EVENTO", "VL_CUSTO"])
    for column in ("MEIO_FINANCEIRO", "CATEGORIA_DESPESA", "COD_PART", "DOC_REF"):
        events[column] = events[column].map(_blank_to_none)
    return events.loc[:, list(EVENT_COLUMNS)]


def _frame_to_entity_profile(frame: pd.DataFrame) -> pd.DataFrame:
    entity_profile = frame.copy()
    for column in ENTITY_PROFILE_COLUMNS:
        entity_profile[column] = entity_profile[column].map(_generic_value_to_text)
    return entity_profile.loc[:, list(ENTITY_PROFILE_COLUMNS)]


def _frame_to_fiscal_event_attributes(frame: pd.DataFrame) -> pd.DataFrame:
    fiscal_event_attributes = frame.copy()
    for column in FISCAL_EVENT_ATTRIBUTE_COLUMNS:
        fiscal_event_attributes[column] = fiscal_event_attributes[column].map(_generic_value_to_text)
    return fiscal_event_attributes.loc[:, list(FISCAL_EVENT_ATTRIBUTE_COLUMNS)]


def _frame_to_tax_scenarios(frame: pd.DataFrame) -> pd.DataFrame:
    tax_scenarios = frame.copy()
    tax_scenarios["DT_REFERENCIA_NORMATIVA"] = tax_scenarios["DT_REFERENCIA_NORMATIVA"].map(_coerce_excel_date)
    tax_scenarios["E_BASELINE"] = pd.Series([_coerce_bool(value) for value in tax_scenarios["E_BASELINE"]], index=tax_scenarios.index, dtype=object)
    tax_scenarios["ATIVO"] = pd.Series([_coerce_bool(value) for value in tax_scenarios["ATIVO"]], index=tax_scenarios.index, dtype=object)
    for column in (
        "ID_CENARIO",
        "ID_ENTIDADE",
        "DESCRICAO",
        "REGIME_ENTIDADE",
        "REGIME_IR",
        "REGIME_CONSUMO",
        "REGIME_ESPECIAL",
        "ID_VERSAO_NORMATIVA",
    ):
        tax_scenarios[column] = tax_scenarios[column].map(_generic_value_to_text)
    return tax_scenarios.loc[:, list(TAX_SCENARIO_COLUMNS)]


def _frame_to_tax_parameters(frame: pd.DataFrame) -> pd.DataFrame:
    tax_parameters = frame.copy()
    for column in ("VIG_INI", "DATA_CONSULTA"):
        tax_parameters[column] = tax_parameters[column].map(_coerce_excel_date)
    tax_parameters["VIG_FIM"] = tax_parameters["VIG_FIM"].map(_coerce_optional_excel_date)
    for column in TAX_PARAMETER_COLUMNS:
        if column == "VALOR":
            tax_parameters[column] = tax_parameters[column].map(_tax_parameter_value_to_text)
        elif column not in {"VIG_INI", "VIG_FIM", "DATA_CONSULTA"}:
            tax_parameters[column] = tax_parameters[column].map(_generic_value_to_text)
    return tax_parameters.loc[:, list(TAX_PARAMETER_COLUMNS)]


def _build_validations(reports_by_stage: Mapping[str, ValidationReport]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for stage, report in reports_by_stage.items():
        if report.ok:
            rows.append(
                {
                    "ETAPA": stage,
                    "OK": True,
                    "ISSUE_CODE": None,
                    "MENSAGEM": "ok",
                    "ACCOUNT_CODE": None,
                    "EVENT_ID": None,
                    "ENTRY_ID": None,
                    "POSTING_ID": None,
                    "ENTITY_ID": None,
                    "SCENARIO_ID": None,
                    "TAX_PARAM_ID": None,
                }
            )
            continue
        for issue in report.issues:
            rows.append(_validation_issue_to_row(stage, issue))
    return pd.DataFrame(rows, columns=VALIDATION_COLUMNS, dtype=object)


def _validation_issue_to_row(stage: str, issue: ValidationIssue) -> dict[str, object]:
    return {
        "ETAPA": stage,
        "OK": False,
        "ISSUE_CODE": issue.code,
        "MENSAGEM": issue.message,
        "ACCOUNT_CODE": issue.account_code,
        "EVENT_ID": issue.event_id,
        "ENTRY_ID": issue.entry_id,
        "POSTING_ID": issue.posting_id,
        "ENTITY_ID": issue.entity_id,
        "SCENARIO_ID": issue.scenario_id,
        "TAX_PARAM_ID": issue.tax_param_id,
    }


def _build_provenance(config: SimulationConfig, events: pd.DataFrame, rule_version: str, tax_context: TaxContext) -> pd.DataFrame:
    event_versions = sorted({str(value) for value in events["SPEC_VERSION"].dropna()})
    tax_normative_versions = sorted({str(value) for value in tax_context.tax_parameters["ID_VERSAO_NORMATIVA"].dropna() if str(value).strip()})
    tax_context_configured = any(
        not frame.empty
        for frame in (
            tax_context.entity_profile,
            tax_context.fiscal_event_attributes,
            tax_context.tax_scenarios,
            tax_context.tax_parameters,
        )
    )
    rows = [
        ("workbook_spec_version", WORKBOOK_SPEC_VERSION),
        ("financial_statement_spec_version", FINANCIAL_STATEMENT_SPEC_VERSION),
        ("tax_interface_spec_version", TAX_INTERFACE_SPEC_VERSION),
        ("simulation_id", config.simulation_id),
        ("scenario_name", config.scenario_name),
        ("simulation_spec_version", config.spec_version),
        ("posting_rule_version", rule_version),
        ("statement_mapping_source", "MAPEAMENTO_DF"),
        ("tax_context_configured", str(tax_context_configured).upper()),
        ("tax_normative_versions", ",".join(tax_normative_versions)),
        ("currency", config.currency),
        ("chart_source", "template_or_input_PLANO_CONTAS"),
        ("event_spec_versions", ",".join(event_versions)),
    ]
    return pd.DataFrame(rows, columns=PROVENANCE_COLUMNS, dtype=object)


def _events_to_workbook(events: pd.DataFrame) -> pd.DataFrame:
    workbook_events = events.copy()
    workbook_events.insert(6, "VL_EVENTO", workbook_events.pop("VL_EVENTO_CENTS").map(_cents_to_excel_money))
    workbook_events.insert(7, "VL_CUSTO", workbook_events.pop("VL_CUSTO_CENTS").map(_optional_cents_to_excel_money))
    return workbook_events.loc[:, list(EVENT_WORKBOOK_COLUMNS)]


def _journal_entries_to_workbook(journal_entries: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(journal_entries, {"VL_LCTO_CENTS": "VL_LCTO"}, JOURNAL_ENTRY_WORKBOOK_COLUMNS)


def _postings_to_workbook(postings: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(postings, {"VL_DC_CENTS": "VL_DC"}, POSTING_WORKBOOK_COLUMNS)


def _journal_to_workbook(journal: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(journal, {"VL_DC_CENTS": "VL_DC"}, JOURNAL_VIEW_WORKBOOK_COLUMNS)


def _ledger_to_workbook(ledger: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(
        ledger,
        {
            "DEBITO_CENTS": "DEBITO",
            "CREDITO_CENTS": "CREDITO",
            "MOVIMENTO_ASSINADO_CENTS": "MOVIMENTO_ASSINADO",
            "SALDO_ASSINADO_CENTS": "SALDO_ASSINADO",
            "SALDO_ABS_CENTS": "SALDO_ABS",
        },
        LEDGER_VIEW_WORKBOOK_COLUMNS,
    )


def _trial_balance_to_workbook(trial_balance: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(
        trial_balance,
        {
            "VL_SLD_INI_CENTS": "VL_SLD_INI",
            "VL_DEB_CENTS": "VL_DEB",
            "VL_CRED_CENTS": "VL_CRED",
            "VL_SLD_FIN_CENTS": "VL_SLD_FIN",
        },
        TRIAL_BALANCE_WORKBOOK_COLUMNS,
    )


def _balance_sheet_to_workbook(balance_sheet: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(balance_sheet, {"VL_CENTS": "VL"}, BALANCE_SHEET_WORKBOOK_COLUMNS)


def _income_statement_to_workbook(income_statement: pd.DataFrame) -> pd.DataFrame:
    return _replace_money_columns(income_statement, {"VL_CENTS": "VL"}, INCOME_STATEMENT_WORKBOOK_COLUMNS)


def _replace_money_columns(
    frame: pd.DataFrame,
    replacements: Mapping[str, str],
    output_columns: tuple[str, ...],
) -> pd.DataFrame:
    workbook_frame = frame.copy()
    for cents_column, workbook_column in replacements.items():
        position = list(workbook_frame.columns).index(cents_column)
        workbook_frame.insert(position, workbook_column, workbook_frame.pop(cents_column).map(_cents_to_excel_money))
    return workbook_frame.loc[:, list(output_columns)]


def _serialize_frame(frame: pd.DataFrame, columns: tuple[str, ...]) -> pd.DataFrame:
    serialized = frame.copy()
    for column in columns:
        if column not in serialized.columns:
            serialized[column] = None
    return serialized.loc[:, list(columns)]


def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ("Workbook contábil parametrizado", None),
        ("Versão", WORKBOOK_SPEC_VERSION),
        ("Regra operacional", "Editar somente abas de entrada; regenerar pelo Python."),
        ("Entradas", "CONFIG, ENTIDADE, PLANO_CONTAS, MAPEAMENTO_CONTAS, EVENTOS, EVENTOS_FISCAIS, MAPEAMENTO_DF, CENARIOS_TRIBUTARIOS e FISCAL_PARAM."),
        ("COD_DF", "PLANO_CONTAS.COD_DF é espelho denormalizado de MAPEAMENTO_DF e é sobrescrito na regeneração."),
        ("Tributário", "Spec 08 materializa contexto tributário contrafactual; não calcula bases, alíquotas, créditos, débitos ou apuração."),
        ("Derivadas", "LANCAMENTOS, PARTIDAS, VINCULO_EVENTO_LCTO, DIARIO, RAZAO, BALANCETE, BP, DRE, VALIDACOES e PROVENIENCIA."),
        ("Fonte de verdade", "Objetos derivados são reconstruídos a partir das abas de entrada."),
    ]
    for row_number, values in enumerate(rows, start=1):
        ws.cell(row=row_number, column=1, value=values[0])
        ws.cell(row=row_number, column=2, value=values[1])
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96


def _write_table(wb: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    materialized = frame.where(pd.notna(frame), None)
    columns = list(materialized.columns)
    ws.append(columns)
    for row in materialized.itertuples(index=False, name=None):
        ws.append(list(row))

    ws.freeze_panes = "A2"
    _style_header(ws, sheet_name)
    _format_columns(ws, columns)
    _set_column_widths(ws, columns)

    table_ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row, 1)}"
    table = Table(displayName=TABLE_NAMES[sheet_name], ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _style_header(ws, sheet_name: str) -> None:
    fill_color = "D9EAF7" if sheet_name in EDITABLE_SHEETS else "E7E6E6"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=fill_color)


def _format_columns(ws, columns: list[str]) -> None:
    money_columns = {
        "VL_EVENTO",
        "VL_CUSTO",
        "VL_LCTO",
        "VL_DC",
        "DEBITO",
        "CREDITO",
        "MOVIMENTO_ASSINADO",
        "SALDO_ASSINADO",
        "SALDO_ABS",
        "VL_SLD_INI",
        "VL_DEB",
        "VL_CRED",
        "VL_SLD_FIN",
        "VL",
    }
    date_columns = {
        "DT_ALT",
        "DT_EVENTO",
        "DT_LCTO",
        "DT_LCTO_EXT",
        "DT_INI",
        "DT_FIN",
        "DT_REF",
        "DT_REFERENCIA_NORMATIVA",
        "VIG_INI",
        "VIG_FIM",
        "DATA_CONSULTA",
    }
    for column_index, column_name in enumerate(columns, start=1):
        number_format = None
        if column_name in money_columns:
            number_format = '#,##0.00'
        elif column_name in date_columns:
            number_format = "yyyy-mm-dd"
        if number_format is None:
            continue
        for cell in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=ws.max_row):
            for item in cell:
                item.number_format = number_format


def _set_column_widths(ws, columns: list[str]) -> None:
    for index, column_name in enumerate(columns, start=1):
        width = min(max(len(column_name) + 2, 12), 34)
        if column_name in {"HIST", "MENSAGEM"}:
            width = 42
        if column_name in {"COD_CTA", "COD_CTA_SUP", "PAPEL_CONTABIL", "COD_LINHA"}:
            width = 22
        ws.column_dimensions[get_column_letter(index)].width = width


def _apply_editable_validations(wb: Workbook) -> None:
    _add_list_validation(wb["CONFIG"], "B", '"BRL"', 5, 5)
    _add_list_validation(wb["PLANO_CONTAS"], "B", '"01,02,03,04,05,09"', 2, 1000)
    _add_list_validation(wb["PLANO_CONTAS"], "C", '"S,A"', 2, 1000)
    _add_list_validation(wb["PLANO_CONTAS"], "H", '"D,C"', 2, 1000)
    _add_list_validation(wb["PLANO_CONTAS"], "J", '"TRUE,FALSE"', 2, 1000)
    _add_list_validation(wb["PLANO_CONTAS"], "K", '"observada,sintética,template,ajustada"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "C", f'"{",".join(item.value for item in EventClass)}"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "D", f'"{",".join(item.value for item in EventType)}"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "E", f'"{",".join(item.value for item in EventDirection)}"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "F", f'"{",".join(item.value for item in EventNature)}"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "I", '"caixa,banco"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "J", '"salarios,aluguel,utilidades,juros"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "L", f'"{",".join(item.value for item in PaymentTerm)}"', 2, 1000)
    _add_list_validation(wb["EVENTOS"], "O", f'"{",".join(item.value for item in Origin)}"', 2, 1000)
    _add_list_validation(wb["MAPEAMENTO_DF"], "B", '"BP,DRE"', 2, 1000)
    _add_list_validation(wb["ENTIDADE"], "D", '"str,int,decimal,bool,date"', 2, 1000)
    _add_list_validation(wb["ENTIDADE"], "E", f'"{",".join(item.value for item in Origin)}"', 2, 1000)
    _add_list_validation(wb["EVENTOS_FISCAIS"], "D", '"str,int,decimal,bool,date"', 2, 1000)
    _add_list_validation(wb["EVENTOS_FISCAIS"], "E", f'"{",".join(item.value for item in Origin)}"', 2, 1000)
    _add_list_validation(wb["CENARIOS_TRIBUTARIOS"], "D", '"TRUE,FALSE"', 2, 1000)
    _add_list_validation(wb["CENARIOS_TRIBUTARIOS"], "K", '"TRUE,FALSE"', 2, 1000)
    _add_list_validation(wb["FISCAL_PARAM"], "G", '"str,int,decimal,bool,date"', 2, 1000)
    _add_list_validation(wb["FISCAL_PARAM"], "H", '"norm,reg,tec,oper"', 2, 1000)


def _add_list_validation(ws, column: str, formula: str, first_row: int, last_row: int) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{column}{first_row}:{column}{last_row}")


def _read_table_frame(ws, expected_columns: tuple[str, ...]) -> pd.DataFrame:
    if not ws.tables:
        raise SchemaValidationError(f"Aba {ws.title} não contém tabela nomeada.")
    table = next(iter(ws.tables.values()))
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        raise SchemaValidationError(f"Tabela vazia em {ws.title}.")
    headers = tuple(str(value).strip() for value in rows[0])
    if headers != expected_columns:
        raise SchemaValidationError(f"Schema físico inválido em {ws.title}: {headers}.")
    data_rows = [row for row in rows[1:] if any(value is not None for value in row)]
    return pd.DataFrame(data_rows, columns=expected_columns, dtype=object)


def _coerce_excel_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    return parse_iso_date(value)  # type: ignore[arg-type]


def _coerce_optional_excel_date(value: object) -> date | None:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    return _coerce_excel_date(value)


def _coerce_int(value: object, field_name: str) -> int:
    if isinstance(value, bool):
        raise SchemaValidationError(f"{field_name} deve ser inteiro, não bool.")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise SchemaValidationError(f"{field_name} inválido: {value!r}.") from exc


def _coerce_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "sim", "s"}:
            return True
        if lowered in {"false", "0", "nao", "não", "n"}:
            return False
    raise SchemaValidationError(f"Valor booleano inválido: {value!r}.")


def _blank_to_none(value: object) -> object | None:
    if pd.isna(value):
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _generic_value_to_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _tax_parameter_value_to_text(value: object) -> str:
    if isinstance(value, float):
        raise SchemaValidationError("FISCAL_PARAM.VALOR deve ser texto; float binário não pode ser fonte de verdade normativa.")
    return _generic_value_to_text(value)


def _copy_tax_context(tax_context: TaxContext) -> TaxContext:
    return TaxContext(
        entity_profile=tax_context.entity_profile.copy(deep=True),
        fiscal_event_attributes=tax_context.fiscal_event_attributes.copy(deep=True),
        tax_scenarios=tax_context.tax_scenarios.copy(deep=True),
        tax_parameters=tax_context.tax_parameters.copy(deep=True),
    )


def _cents_to_excel_money(value: int) -> Decimal:
    return Decimal(int(value)) / Decimal("100")


def _optional_cents_to_excel_money(value: object) -> Decimal | None:
    if pd.isna(value) or value is None:
        return None
    return _cents_to_excel_money(int(value))


def _excel_money_to_cents(value: object, *, optional: bool) -> int | None:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        if optional:
            return None
        raise SchemaValidationError("Valor monetário obrigatório ausente.")
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise SchemaValidationError(f"Valor monetário inválido: {value!r}.") from exc
    cents = decimal_value * Decimal("100")
    if cents != cents.to_integral_value():
        raise SchemaValidationError("Valor monetário deve ter no máximo duas casas decimais.")
    return int(cents)


def _raise_if_invalid(stage: str, report: ValidationReport) -> None:
    if report.ok:
        return
    details = "; ".join(issue.code for issue in report.issues[:5])
    raise AccountingInvariantError(f"{stage} inválido: {details}")
